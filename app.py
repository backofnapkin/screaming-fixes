"""
Screaming Fixes v3 - AI-Powered WordPress Broken Link Fixer
Built with Streamlit

Upload your Screaming Frog CSV export → Review unique broken URLs →
Get AI suggestions → Approve fixes → Apply to WordPress

Supports:
- Broken Links (4xx errors)
- Redirect Chains
- Image Alt Text optimization
"""

import os
import io
import json
import re
import time
from datetime import datetime
from typing import Optional, Dict, List, Any
from urllib.parse import urlparse
from collections import defaultdict

import streamlit as st
import pandas as pd

# Import configuration from centralized config
from config import (
    QUICK_START_API_KEY,
    QUICK_START_FREE_SUGGESTIONS,
    QUICK_START_PAGE_LIMIT,
    AGENT_MODE_API_KEY,
    AGENT_MODE_FREE_SUGGESTIONS,
    AGENT_MODE_LIMIT,
    LANGSMITH_ENABLED,
    get_app_css,
)

# Import session state initialization
from utils.session_state import init_session_state

# Import Claude API service
from services.claude_api import (
    get_ai_suggestion,
    get_ai_alt_text_suggestion,
    track_event,
    is_anthropic_available,
)

# Import sidebar/integrations components
from components.sidebar import (
    get_integration_status,
    render_feature_cards,
    render_integrations_panel,
)

# Import backlink reclaim feature
from features.backlink_reclaim import (
    init_backlink_reclaim_state,
    reset_backlink_reclaim_state,
    load_scan_results,
    render_backlink_fix_workflow,
)

# Import DataForSEO client
from services.dataforseo_api import DataForSEOClient

# Page config
st.set_page_config(
    page_title="Screaming Fixes",
    page_icon="🔧",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# Optional imports - Anthropic is now handled by services/claude_api.py
ANTHROPIC_AVAILABLE = is_anthropic_available()

try:
    from wordpress_client import WordPressClient
    WP_AVAILABLE = True
except ImportError:
    WP_AVAILABLE = False


# =============================================================================
# APPLY CSS
# =============================================================================

st.markdown(get_app_css(), unsafe_allow_html=True)


# =============================================================================
# DATA PROCESSING
# =============================================================================

def detect_domain(urls: List[str]) -> Optional[str]:
    """Detect primary domain from URLs"""
    counts = defaultdict(int)
    for url in urls:
        try:
            parsed = urlparse(url)
            domain = parsed.netloc.lower().replace('www.', '')
            if domain:
                counts[domain] += 1
        except:
            pass
    return max(counts, key=counts.get) if counts else None


def is_internal(url: str, domain: str) -> bool:
    """Check if URL is internal"""
    if not domain:
        return True
    try:
        parsed = urlparse(url)
        url_domain = parsed.netloc.lower().replace('www.', '')
        return domain in url_domain or url_domain in domain
    except:
        return False


def parse_csv(uploaded_file) -> Optional[pd.DataFrame]:
    """Parse uploaded CSV"""
    try:
        df = pd.read_csv(uploaded_file)
        df.columns = df.columns.str.strip()
        
        required = ['Source', 'Destination', 'Status Code']
        missing = [c for c in required if c not in df.columns]
        if missing:
            st.error(f"Missing columns: {', '.join(missing)}")
            return None
        
        # Check for post_id column (case-insensitive)
        post_id_col = None
        for col in df.columns:
            if col.lower().replace('_', '').replace(' ', '') in ['postid', 'post_id', 'id']:
                post_id_col = col
                break
        
        if post_id_col:
            df['post_id'] = pd.to_numeric(df[post_id_col], errors='coerce')
            valid_ids = df['post_id'].notna().sum()
            st.success(f"✅ Found Post ID column: **{valid_ids}** valid IDs detected")
        else:
            df['post_id'] = None
        
        if 'Link Position' in df.columns:
            before = len(df)
            df = df[df['Link Position'] == 'Content'].copy()
            filtered = before - len(df)
            if filtered > 0:
                st.info(f"📍 Filtered to Content links only ({filtered} non-content links excluded for safety)")
        
        return df
    except Exception as e:
        st.error(f"Error parsing CSV: {e}")
        return None


def detect_csv_type(df: pd.DataFrame) -> str:
    """
    Auto-detect CSV type based on columns.
    Returns: 'post_ids', 'redirect_chains', 'image_alt_text', or 'broken_links'
    """
    columns = set(df.columns.str.lower().str.strip())
    
    # Check for Post ID file first
    # Post ID files have Address + post_id column, but NO Destination or Final Address
    has_address = 'address' in columns
    has_post_id = any(col.startswith('post_id') or col.startswith('postid') or col == 'post-id' 
                      for col in columns)
    has_destination = 'destination' in columns
    has_final_address = 'final address' in columns
    
    # Post ID file: has Address and post_id, but NOT a report file
    if has_address and has_post_id and not has_destination and not has_final_address:
        return 'post_ids'
    
    # Redirect chains has these distinctive columns
    redirect_chain_indicators = {'final address', 'number of redirects', 'chain type', 'loop'}
    
    # Check for redirect chains first (more specific)
    if len(redirect_chain_indicators & columns) >= 2:
        return 'redirect_chains'
    
    # Broken links has Status Code - this is the key differentiator
    has_status_code = 'status code' in columns
    has_source = 'source' in columns
    has_type = 'type' in columns
    has_alt_text = 'alt text' in columns
    
    # If it has Status Code, it's likely broken links (4xx errors)
    # Check the status codes to confirm - broken links have 4xx/5xx status codes
    if has_status_code and has_source and has_destination:
        # Check actual status code values
        status_col = df.columns[df.columns.str.lower().str.strip() == 'status code'][0]
        try:
            status_values = pd.to_numeric(df[status_col], errors='coerce').dropna()
            if len(status_values) > 0:
                # If we have 4xx or 5xx status codes, it's a broken links report
                avg_status = status_values.mean()
                if avg_status >= 400:
                    return 'broken_links'
        except:
            pass
    
    # Image Alt Text has these distinctive columns (from All Image Inlinks export)
    # Key differentiator: Type column with "Image" values (not just "Hyperlink")
    # AND typically no error status codes
    if has_type and has_alt_text and has_source and has_destination:
        # Check actual values in Type column to confirm it's an image report
        type_col = df.columns[df.columns.str.lower().str.strip() == 'type'][0]
        type_values = df[type_col].astype(str).str.lower().unique()
        # Image reports have "Image" type - "Hyperlink" alone is NOT enough
        if 'image' in type_values:
            return 'image_alt_text'
    
    # Fall back to broken links if has the basic columns
    broken_link_indicators = {'destination', 'status code'}
    if len(broken_link_indicators & columns) >= 2:
        return 'broken_links'
    
    # Default to broken links if unclear
    return 'broken_links'


def parse_redirect_chains_csv(uploaded_file) -> Optional[pd.DataFrame]:
    """Parse uploaded Redirect Chains CSV from Screaming Frog"""
    try:
        df = pd.read_csv(uploaded_file)
        df.columns = df.columns.str.strip()
        
        # Required columns for redirect chains
        required = ['Source', 'Address', 'Final Address']
        missing = [c for c in required if c not in df.columns]
        if missing:
            st.error(f"Missing columns: {', '.join(missing)}")
            return None
        
        # Ensure we have the key columns with defaults
        if 'Number of Redirects' not in df.columns:
            df['Number of Redirects'] = 1
        if 'Loop' not in df.columns:
            df['Loop'] = False
        if 'Link Position' not in df.columns:
            df['Link Position'] = 'Content'
        if 'Temp Redirect in Chain' not in df.columns:
            df['Temp Redirect in Chain'] = False
        if 'Anchor Text' not in df.columns:
            df['Anchor Text'] = ''
        
        # Convert Loop column to boolean
        df['Loop'] = df['Loop'].astype(str).str.upper() == 'TRUE'
        
        # Convert Temp Redirect to boolean  
        df['Temp Redirect in Chain'] = df['Temp Redirect in Chain'].astype(str).str.upper() == 'TRUE'
        
        # Check for post_id column (case-insensitive)
        post_id_col = None
        for col in df.columns:
            if col.lower().replace('_', '').replace(' ', '') in ['postid', 'post_id', 'id']:
                post_id_col = col
                break
        
        if post_id_col:
            df['post_id'] = pd.to_numeric(df[post_id_col], errors='coerce')
            valid_ids = df['post_id'].notna().sum()
            st.success(f"✅ Found Post ID column: **{valid_ids}** valid IDs detected")
        else:
            df['post_id'] = None
        
        return df
    except Exception as e:
        st.error(f"Error parsing CSV: {e}")
        return None


def group_redirect_chains(df: pd.DataFrame, domain: str) -> tuple:
    """
    Group redirect chains by unique Address -> Final Address pairs.
    Also separates out sitewide links and loops.
    
    Returns: (redirects_dict, sitewide_list, loops_list)
    """
    redirects = {}
    sitewide = []
    loops = []
    
    for _, row in df.iterrows():
        address = row['Address']
        final_address = row['Final Address']
        source = row['Source']
        link_position = row.get('Link Position', 'Content')
        is_loop = row.get('Loop', False)
        is_temp = row.get('Temp Redirect in Chain', False)
        num_hops = int(row.get('Number of Redirects', 1))
        anchor = row.get('Anchor Text', '')
        
        # Handle loops separately
        if is_loop:
            loops.append({
                'address': address,
                'final_address': final_address,
                'source': source,
                'anchor': anchor,
                'hops': num_hops
            })
            continue
        
        # Handle sitewide (non-Content) links separately
        if link_position and link_position.lower() not in ['content', '']:
            sitewide.append({
                'address': address,
                'final_address': final_address,
                'source': source,
                'position': link_position,
                'anchor': anchor
            })
            continue
        
        # Create unique key for this redirect pair
        key = f"{address}|||{final_address}"
        
        if key not in redirects:
            redirects[key] = {
                'address': address,
                'final_address': final_address,
                'is_internal': is_internal(address, domain),
                'is_temp_redirect': is_temp,
                'num_hops': num_hops,
                'anchors': [],
                'sources': [],
                'source_post_ids': {},
                'count': 0
            }
        
        redirects[key]['count'] += 1
        
        if anchor and anchor not in redirects[key]['anchors']:
            redirects[key]['anchors'].append(anchor)
        
        if source not in redirects[key]['sources']:
            redirects[key]['sources'].append(source)
        
        # Store post_id if available
        if pd.notna(row.get('post_id')):
            redirects[key]['source_post_ids'][source] = int(row['post_id'])
        
        # Update temp redirect status if any in chain is temp
        if is_temp:
            redirects[key]['is_temp_redirect'] = True
    
    # Consolidate sitewide links by address
    sitewide_consolidated = {}
    for item in sitewide:
        key = item['address']
        if key not in sitewide_consolidated:
            sitewide_consolidated[key] = {
                'address': item['address'],
                'final_address': item['final_address'],
                'position': item['position'],
                'sources': [],
                'count': 0
            }
        sitewide_consolidated[key]['sources'].append(item['source'])
        sitewide_consolidated[key]['count'] += 1
    
    # Consolidate loops by address
    loops_consolidated = {}
    for item in loops:
        key = item['address']
        if key not in loops_consolidated:
            loops_consolidated[key] = {
                'address': item['address'],
                'final_address': item['final_address'],
                'sources': [],
                'count': 0
            }
        loops_consolidated[key]['sources'].append(item['source'])
        loops_consolidated[key]['count'] += 1
    
    return redirects, list(sitewide_consolidated.values()), list(loops_consolidated.values())


# =============================================================================
# IMAGE ALT TEXT PARSING
# =============================================================================

def parse_image_alt_text_csv(uploaded_file) -> Optional[pd.DataFrame]:
    """
    Parse uploaded All Image Inlinks CSV from Screaming Frog.
    Filters to Content position only and identifies images needing alt text fixes.
    """
    try:
        df = pd.read_csv(uploaded_file)
        df.columns = df.columns.str.strip()
        
        # Required columns
        required = ['Source', 'Destination', 'Alt Text']
        missing = [c for c in required if c not in df.columns]
        if missing:
            st.error(f"Missing columns: {', '.join(missing)}")
            return None
        
        # Get Link Position if available for filtering
        if 'Link Position' in df.columns:
            before = len(df)
            # Keep only Content position images
            df = df[df['Link Position'].str.lower().isin(['content', ''])].copy()
            filtered_pos = before - len(df)
            if filtered_pos > 0:
                st.info(f"📍 Filtered to Content images only ({filtered_pos} header/footer/sidebar images excluded)")
        
        # Check for post_id column (case-insensitive)
        post_id_col = None
        for col in df.columns:
            if col.lower().replace('_', '').replace(' ', '') in ['postid', 'post_id', 'id']:
                post_id_col = col
                break
        
        if post_id_col:
            df['post_id'] = pd.to_numeric(df[post_id_col], errors='coerce')
            valid_ids = df['post_id'].notna().sum()
            st.success(f"✅ Found Post ID column: **{valid_ids}** valid IDs detected")
        else:
            df['post_id'] = None
        
        return df
    except Exception as e:
        st.error(f"Error parsing CSV: {e}")
        return None


def is_excluded_image(img_url: str) -> bool:
    """Check if image URL matches exclusion patterns (logos, icons, etc.)"""
    img_lower = img_url.lower()
    
    exclusion_patterns = [
        'logo',
        'icon',
        'favicon',
        'sprite',
        'placeholder',
        'avatar',
        'gravatar.com',
        'badge',
        'button',
        'social',
        'facebook',
        'twitter',
        'linkedin',
        'instagram',
        'youtube',
        'pinterest',
        'background',
        'bg-',
        '-bg.',
    ]
    
    return any(pattern in img_lower for pattern in exclusion_patterns)


def is_excluded_page(source_url: str) -> bool:
    """Check if source URL is a dynamic/listing page that should be excluded"""
    source_lower = source_url.lower()
    
    # Exact homepage match
    parsed = urlparse(source_url)
    if parsed.path in ['', '/', '/index.html', '/index.php']:
        return True
    
    exclusion_patterns = [
        '/page/',
        '/category/',
        '/tag/',
        '/author/',
        '?listing-page=',
        '?paged=',
        '/wp-admin/',
    ]
    
    return any(pattern in source_lower for pattern in exclusion_patterns)


def is_bad_alt_text(alt_text: str) -> tuple:
    """
    Check if alt text is missing or non-descriptive.
    Returns: (is_bad: bool, reason: str)
    """
    if pd.isna(alt_text) or alt_text.strip() == '':
        return True, 'missing'
    
    alt = alt_text.strip()
    
    # Filename patterns
    filename_patterns = [
        r'^IMG_\d+',           # IMG_0369
        r'^DSC[_\d]+',         # DSC_0042, DSC0042
        r'^DCIM',              # DCIM photos
        r'^Photo\d*',          # Photo1, Photo
        r'^Image[-_]?\d*',     # Image-1, Image_1, Image1
        r'^pic\d+',            # pic1, pic2
        r'^screenshot',        # screenshot-2024-01-15
        r'^screen[-_]?shot',   # Screen-Shot, Screen_Shot
        r'^\d{6,}',            # Long numeric strings like 556316_444422658962091
        r'^[A-F0-9]{8}-[A-F0-9]{4}',  # UUID patterns
        r'^\d+[-_]\d+',        # Patterns like 308395697_429013265790380
    ]
    
    for pattern in filename_patterns:
        if re.match(pattern, alt, re.IGNORECASE):
            return True, 'filename'
    
    # Too short to be descriptive (less than 5 chars, excluding common words)
    if len(alt) < 5 and alt.lower() not in ['logo', 'icon', 'menu', 'back', 'next']:
        return True, 'too_short'
    
    return False, 'ok'


def group_images_for_alt_text(df: pd.DataFrame, domain: str) -> tuple:
    """
    Group images by Destination (image URL) for alt text fixing.
    Filters out excluded images and pages.
    
    Returns: (images_dict, excluded_count)
    """
    images = {}
    excluded_count = 0
    
    for _, row in df.iterrows():
        source = row['Source']
        destination = row['Destination']
        alt_text = row.get('Alt Text', '')
        img_type = row.get('Type', 'Image')
        
        # Skip excluded pages
        if is_excluded_page(source):
            excluded_count += 1
            continue
        
        # Skip excluded image types
        if is_excluded_image(destination):
            excluded_count += 1
            continue
        
        # Check if alt text needs fixing
        is_bad, reason = is_bad_alt_text(alt_text)
        if not is_bad:
            excluded_count += 1
            continue
        
        # Use destination (image URL) as the key
        key = destination
        
        if key not in images:
            images[key] = {
                'image_url': destination,
                'current_alt': alt_text if pd.notna(alt_text) else '',
                'alt_status': reason,  # 'missing', 'filename', 'too_short'
                'img_type': img_type,  # 'Image' or 'Hyperlink'
                'sources': [],
                'source_post_ids': {},
                'count': 0
            }
        
        images[key]['count'] += 1
        
        if source not in images[key]['sources']:
            images[key]['sources'].append(source)
        
        # Store post_id if available
        if pd.notna(row.get('post_id')):
            images[key]['source_post_ids'][source] = int(row['post_id'])
    
    return images, excluded_count


def parse_post_id_csv(uploaded_file) -> Dict[str, int]:
    """
    Parse a Custom Extraction CSV containing Post IDs.
    Returns a dict mapping URL/Address to Post ID.
    
    Handles Screaming Frog's column naming conventions:
    - 'post_id 1', 'post_id 2' (numbered extractors)
    - 'post_id', 'PostId', 'post-id' (various formats)
    """
    try:
        df = pd.read_csv(uploaded_file)
        df.columns = df.columns.str.strip()
        
        # Find the URL column (could be 'Address', 'URL', or similar)
        url_col = None
        for col in df.columns:
            if col.lower() in ['address', 'url', 'source', 'page url', 'page']:
                url_col = col
                break
        
        if not url_col:
            # Default to first column if no match
            url_col = df.columns[0]
        
        # Find the Post ID column - handle various naming conventions
        post_id_col = None
        for col in df.columns:
            col_lower = col.lower()
            # Check for exact matches first
            if col_lower in ['post_id', 'postid', 'post-id', 'id', 'page_id', 'pageid']:
                post_id_col = col
                break
            # Check for Screaming Frog's numbered format: 'post_id 1', 'post_id 2', etc.
            if col_lower.startswith('post_id') or col_lower.startswith('postid') or col_lower.startswith('post-id'):
                post_id_col = col
                break
            # Also check for 'page_id 1' format
            if col_lower.startswith('page_id') or col_lower.startswith('pageid'):
                post_id_col = col
                break
        
        if not post_id_col:
            return {}
        
        # Build the mapping
        post_id_map = {}
        for _, row in df.iterrows():
            url = row[url_col]
            post_id = row[post_id_col]
            
            if pd.notna(url) and pd.notna(post_id):
                try:
                    # Handle potential float values from CSV
                    post_id_map[str(url).strip()] = int(float(post_id))
                except (ValueError, TypeError):
                    continue
        
        return post_id_map
    
    except Exception as e:
        st.error(f"Error parsing Post ID CSV: {e}")
        return {}


def match_post_ids_to_sources(source_urls: List[str]) -> tuple:
    """
    Match source URLs from a report to the Post ID cache.
    Returns (matched_count, unmatched_urls)
    """
    matched = 0
    unmatched = []
    
    for url in source_urls:
        if url in st.session_state.post_id_cache:
            matched += 1
        else:
            # Check if it's an archive/category page (expected to be unmatched)
            url_lower = url.lower()
            is_archive = any(x in url_lower for x in ['/category/', '/tag/', '/author/', '/page/', '/archive/'])
            unmatched.append({
                'url': url,
                'is_archive': is_archive
            })
    
    return matched, unmatched


def get_unmatched_summary(unmatched_urls: List[Dict]) -> Dict:
    """Summarize unmatched URLs by type"""
    archive_count = sum(1 for u in unmatched_urls if u['is_archive'])
    other_count = len(unmatched_urls) - archive_count
    
    return {
        'total': len(unmatched_urls),
        'archive': archive_count,
        'other': other_count
    }


def group_by_broken_url(df: pd.DataFrame, domain: str) -> Dict[str, Dict]:
    """Group data by unique broken URL"""
    grouped = {}
    
    for _, row in df.iterrows():
        dest = row['Destination']
        
        if dest not in grouped:
            grouped[dest] = {
                'status_code': row['Status Code'],
                'status_text': row.get('Status', ''),
                'is_internal': is_internal(dest, domain),
                'anchors': [],
                'sources': [],
                'source_post_ids': {},  # Map source URL to post_id
                'count': 0
            }
        
        grouped[dest]['count'] += 1
        
        anchor = row.get('Anchor', '')
        if anchor and anchor not in grouped[dest]['anchors']:
            grouped[dest]['anchors'].append(anchor)
        
        source = row['Source']
        if source not in grouped[dest]['sources']:
            grouped[dest]['sources'].append(source)
        
        # Store post_id if available
        if pd.notna(row.get('post_id')):
            grouped[dest]['source_post_ids'][source] = int(row['post_id'])
    
    return grouped


# =============================================================================
# UI COMPONENTS
# =============================================================================

def render_nav():
    """Render top navigation bar"""
    st.markdown("""
    <div style="display: flex; justify-content: space-between; align-items: center; padding: 0.75rem 0; border-bottom: 1px solid #99f6e4; margin-bottom: 1.5rem;">
        <div style="font-size: 1.5rem; font-weight: 700; color: #0d9488;">🔧 Screaming Fixes</div>
        <div style="display: flex; gap: 2rem;">
            <a href="#" style="color: #0d9488; text-decoration: none; font-weight: 500; font-size: 0.95rem;">Home</a>
            <a href="#about" style="color: #64748b; text-decoration: none; font-weight: 500; font-size: 0.95rem;">About</a>
            <a href="#instructions" style="color: #64748b; text-decoration: none; font-weight: 500; font-size: 0.95rem;">Instructions</a>
            <a href="mailto:brett.lindenberg@gmail.com" style="color: #64748b; text-decoration: none; font-weight: 500; font-size: 0.95rem;">Contact</a>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_header():
    """Render the header and introduction"""
    st.markdown("""
    <div style="background: linear-gradient(135deg, #f0fdfa 0%, #ecfeff 50%, #f0f9ff 100%); 
                padding: 1.75rem 2rem; 
                border-radius: 16px; 
                margin: 0 0 1.5rem 0;
                border: 1px solid #99f6e4;">
        <p style="font-size: 1.15rem; line-height: 1.75; color: #134e4a; margin: 0;">
            <strong style="color: #0d9488;">If you're an SEO or manage WordPress websites, you already use Screaming Frog.</strong><br><br>
            But running audits is the easy part. <em>Actually fixing</em> hundreds of broken links and redirect chains? That's where hours disappear.<br><br>
            <strong style="color: #0d9488;">Work smarter.</strong> Upload your CSV file directly from Screaming Frog, review each broken link, and decide: remove it, replace it, or let AI suggest a fix.<br><br>
            Then the real power kicks in — Screaming Fixes connects to your WordPress site and applies every approved fix automatically. No more logging into each post. No more clicking publish.<br><br>
            This isn't just another audit tool. <strong style="color: #0d9488;">This one actually gets the work done.</strong> Export a CSV or JSON to share your updates with clients.
        </p>
    </div>
    """, unsafe_allow_html=True)


# get_integration_status moved to components/sidebar.py


# render_feature_cards moved to components/sidebar.py


# render_integrations_panel moved to components/sidebar.py


def render_upload_section():
    """Render the CSV upload section - always visible with file status cards"""
    st.markdown('<p class="section-header">📤 Upload Reports</p>', unsafe_allow_html=True)
    
    # Check current state
    has_post_ids = st.session_state.post_id_file_uploaded or st.session_state.has_post_ids
    post_id_count = len(st.session_state.post_id_cache)
    has_broken_links = st.session_state.df is not None
    has_redirect_chains = st.session_state.rc_df is not None
    has_image_alt_text = st.session_state.iat_df is not None
    
    st.markdown("""
    Export a report from Screaming Frog and upload it here. We'll auto-detect the file type.
    """)
    
    with st.expander("🔗 Broken Links — How to export", expanded=False):
        st.markdown("""
        1. Run a crawl in Screaming Frog
        2. Go to **Bulk Export → Response Codes → Client Error (4xx) → Inlinks**
        3. Save and upload the CSV
        """)
    
    with st.expander("🔄 Redirect Chains — How to export", expanded=False):
        st.markdown("""
        1. Run a crawl in Screaming Frog
        2. Go to **Reports → Redirects → All Redirects**
        3. Save and upload the CSV
        """)
    
    with st.expander("🖼️ Image Alt Text — How to export", expanded=False):
        st.markdown("""
        1. Run a crawl in Screaming Frog
        2. Go to **Bulk Export → Images → All Image Inlinks**
        3. Save and upload the CSV

        *Requires all integrations to be connected (Post IDs + AI + WordPress)*
        """)

    # Backlink Reclaim scan section
    with st.expander("🔙 Backlink Reclaim — Scan for broken backlinks", expanded=False):
        st.markdown("""
        Scan any domain to find external sites linking to broken pages on your site.
        Fix these by creating redirects to reclaim lost link equity.
        """)

        br_domain_input = st.text_input(
            "Domain to scan",
            placeholder="example.com",
            key="br_domain_input",
            help="Enter your domain without https:// or www."
        )

        col1, col2 = st.columns([1, 3])
        with col1:
            scan_clicked = st.button(
                "🔍 Scan Domain",
                key="br_scan_button",
                disabled=not br_domain_input,
                type="primary"
            )

        if scan_clicked and br_domain_input:
            # Initialize backlink reclaim state if needed
            init_backlink_reclaim_state()

            with st.spinner("Scanning backlinks... (this may take 30 seconds)"):
                try:
                    client = DataForSEOClient()
                    backlinks, broken_count, total_count, api_cost, error = client.get_broken_backlinks(br_domain_input)

                    if error:
                        st.error(f"Scan failed: {error}")
                    elif broken_count == 0:
                        st.info(f"No broken backlinks found for {br_domain_input}. Great news!")
                    else:
                        # Build scan results object matching expected format
                        scan_results = {
                            'domain': br_domain_input,
                            'broken_backlinks': backlinks,
                            'broken_count': broken_count,
                            'total_count': total_count,
                            'api_cost_cents': api_cost,
                        }

                        # Load into backlink reclaim state
                        load_scan_results(scan_results, br_domain_input)

                        # Switch to backlink_reclaim task
                        st.session_state.current_task = 'backlink_reclaim'
                        st.session_state.task_type = 'backlink_reclaim'

                        st.success(f"Found {broken_count} broken backlinks on {br_domain_input}!")
                        st.rerun()
                except Exception as e:
                    st.error(f"Error scanning domain: {str(e)}")

        # Note about DataForSEO
        st.caption("Powered by DataForSEO API. Uses demo data if no API credentials are configured.")

    uploaded = st.file_uploader(
        "Upload your Screaming Frog CSV",
        type=['csv'],
        label_visibility='collapsed',
        key="main_uploader",
        help="Supported: Broken Links, Redirect Chains, Image Alt Text, or Post IDs"
    )
    
    # Show uploaded file status cards
    if has_post_ids or has_broken_links or has_redirect_chains or has_image_alt_text:
        st.markdown("**Uploaded Files:**")
        
        # Post IDs card (shows if uploaded via main uploader)
        if has_post_ids:
            col1, col2 = st.columns([5, 1])
            with col1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%); padding: 0.75rem 1rem; border-radius: 8px; border: 1px solid #6ee7b7; margin-bottom: 0.5rem;">
                    <span style="font-weight: 600; color: #065f46;">🆔 Post IDs</span>
                    <span style="color: #047857; margin-left: 0.5rem;">{post_id_count} URLs mapped — Full Mode enabled</span>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                if st.button("✕", key="clear_post_ids_upload", help="Clear Post IDs"):
                    st.session_state.post_id_cache = {}
                    st.session_state.post_id_file_uploaded = False
                    st.session_state.post_id_file_count = 0
                    st.session_state.has_post_ids = False
                    st.session_state.selected_mode = 'quick_start'
                    st.rerun()
        
        # Broken Links card
        if has_broken_links:
            broken_count = len(st.session_state.broken_urls)
            source_count = st.session_state.source_pages_count
            col1, col2 = st.columns([5, 1])
            with col1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%); padding: 0.75rem 1rem; border-radius: 8px; border: 1px solid #fcd34d; margin-bottom: 0.5rem;">
                    <span style="font-weight: 600; color: #92400e;">🔗 Broken Links</span>
                    <span style="color: #a16207; margin-left: 0.5rem;">{broken_count} unique broken URLs across {source_count} pages</span>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                if st.button("✕", key="clear_broken_links", help="Clear Broken Links"):
                    st.session_state.df = None
                    st.session_state.broken_urls = {}
                    st.session_state.decisions = {}
                    st.session_state.domain = None
                    if st.session_state.current_task == 'broken_links':
                        st.session_state.current_task = 'redirect_chains' if has_redirect_chains else ('image_alt_text' if has_image_alt_text else None)
                        st.session_state.task_type = st.session_state.current_task
                    st.rerun()
        
        # Redirect Chains card
        if has_redirect_chains:
            redirect_count = len(st.session_state.rc_redirects)
            rc_source_count = len(st.session_state.rc_df['Source'].unique()) if st.session_state.rc_df is not None else 0
            col1, col2 = st.columns([5, 1])
            with col1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #e0e7ff 0%, #c7d2fe 100%); padding: 0.75rem 1rem; border-radius: 8px; border: 1px solid #a5b4fc; margin-bottom: 0.5rem;">
                    <span style="font-weight: 600; color: #3730a3;">🔄 Redirect Chains</span>
                    <span style="color: #4338ca; margin-left: 0.5rem;">{redirect_count} unique redirects across {rc_source_count} pages</span>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                if st.button("✕", key="clear_redirect_chains", help="Clear Redirect Chains"):
                    st.session_state.rc_df = None
                    st.session_state.rc_redirects = {}
                    st.session_state.rc_decisions = {}
                    st.session_state.rc_domain = None
                    st.session_state.rc_sitewide = []
                    st.session_state.rc_loops = []
                    if st.session_state.current_task == 'redirect_chains':
                        st.session_state.current_task = 'broken_links' if has_broken_links else ('image_alt_text' if has_image_alt_text else None)
                        st.session_state.task_type = st.session_state.current_task
                    st.rerun()
        
        # Image Alt Text card
        if has_image_alt_text:
            image_count = len(st.session_state.iat_images)
            excluded_count = st.session_state.iat_excluded_count
            # Count unique source pages
            all_sources = set()
            for img_data in st.session_state.iat_images.values():
                all_sources.update(img_data['sources'])
            iat_source_count = len(all_sources)
            col1, col2 = st.columns([5, 1])
            with col1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #fce7f3 0%, #fbcfe8 100%); padding: 0.75rem 1rem; border-radius: 8px; border: 1px solid #f9a8d4; margin-bottom: 0.5rem;">
                    <span style="font-weight: 600; color: #9d174d;">🖼️ Image Alt Text</span>
                    <span style="color: #be185d; margin-left: 0.5rem;">{image_count} images need alt text across {iat_source_count} pages</span>
                    <span style="color: #9ca3af; margin-left: 0.5rem; font-size: 0.85rem;">({excluded_count} filtered out)</span>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                if st.button("✕", key="clear_image_alt_text", help="Clear Image Alt Text"):
                    st.session_state.iat_df = None
                    st.session_state.iat_images = {}
                    st.session_state.iat_decisions = {}
                    st.session_state.iat_domain = None
                    st.session_state.iat_excluded_count = 0
                    if st.session_state.current_task == 'image_alt_text':
                        st.session_state.current_task = 'broken_links' if has_broken_links else ('redirect_chains' if has_redirect_chains else None)
                        st.session_state.task_type = st.session_state.current_task
                    st.rerun()

        # Backlink Reclaim card
        has_backlink_reclaim = st.session_state.br_grouped_pages and len(st.session_state.br_grouped_pages) > 0
        if has_backlink_reclaim:
            br_count = len(st.session_state.br_grouped_pages)
            br_domain = st.session_state.br_domain
            col1, col2 = st.columns([5, 1])
            with col1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #ccfbf1 0%, #99f6e4 100%); padding: 0.75rem 1rem; border-radius: 8px; border: 1px solid #5eead4; margin-bottom: 0.5rem;">
                    <span style="font-weight: 600; color: #0f766e;">🔙 Backlink Reclaim</span>
                    <span style="color: #0d9488; margin-left: 0.5rem;">{br_count} broken backlinks on {br_domain}</span>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                if st.button("✕", key="clear_backlink_reclaim", help="Clear Backlink Reclaim"):
                    reset_backlink_reclaim_state()
                    if st.session_state.current_task == 'backlink_reclaim':
                        st.session_state.current_task = 'broken_links' if has_broken_links else ('redirect_chains' if has_redirect_chains else ('image_alt_text' if has_image_alt_text else None))
                        st.session_state.task_type = st.session_state.current_task
                    st.rerun()

        st.markdown("")  # Spacing
    
    # Privacy and disclaimer (only show if no files uploaded yet)
    if not has_post_ids and not has_broken_links and not has_redirect_chains and not has_image_alt_text:
        st.markdown("""
        <div class="privacy-notice" style="margin-top: 0.5rem;">
            🔒 <strong>Your data is safe</strong> — processed in your browser session only. Nothing is saved to any database.
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div style="font-size: 0.8rem; color: #64748b; padding: 0.75rem; background: #f8fafc; border-radius: 8px; margin-top: 0.5rem;">
            <strong>Disclaimer:</strong> You are responsible for reviewing and approving all changes before they are applied to your site. 
            Screaming Frog data may occasionally contain errors or false positives. Always verify fixes are appropriate for your specific situation.
        </div>
        """, unsafe_allow_html=True)
    
    # Process new upload
    if uploaded:
        # Only process if we don't already have this data loaded
        # (prevents re-processing on every rerun while file is still in uploader)
        already_processed = False
        
        # Check if this file was already processed by comparing basic stats
        df_preview = pd.read_csv(uploaded)
        uploaded.seek(0)  # Reset file pointer
        
        csv_type = detect_csv_type(df_preview)
        
        # Determine if we should process based on what's already loaded
        if csv_type == 'post_ids':
            # Only process if we don't have post IDs or the count is different
            already_processed = has_post_ids
        elif csv_type == 'redirect_chains':
            already_processed = has_redirect_chains
        elif csv_type == 'image_alt_text':
            already_processed = st.session_state.iat_df is not None
        else:  # broken_links
            already_processed = has_broken_links
        
        if not already_processed:
            success = False
            if csv_type == 'post_ids':
                success = process_post_id_upload(uploaded)
            elif csv_type == 'redirect_chains':
                success = process_redirect_chains_upload(uploaded)
            elif csv_type == 'image_alt_text':
                success = process_image_alt_text_upload(uploaded)
            else:
                success = process_broken_links_upload(uploaded)
            
            if success:
                st.rerun()


def process_post_id_upload(uploaded_file) -> bool:
    """Process an uploaded Post ID CSV (Custom Extraction from Screaming Frog). Returns True if processing succeeded."""
    post_id_map = parse_post_id_csv(uploaded_file)
    
    if post_id_map:
        # Store in session state
        st.session_state.post_id_cache.update(post_id_map)
        st.session_state.post_id_file_uploaded = True
        st.session_state.post_id_file_count = len(post_id_map)
        st.session_state.has_post_ids = True
        st.session_state.selected_mode = 'full'
        
        # Track upload
        track_event("csv_upload", {
            "type": "post_ids",
            "post_id_count": len(post_id_map)
        })
        
        # Retroactive matching if report already loaded
        if st.session_state.df is not None:
            source_urls = st.session_state.df['Source'].unique().tolist()
            matched, unmatched = match_post_ids_to_sources(source_urls)
            st.session_state.post_id_matched_count = matched
            st.session_state.post_id_unmatched_urls = unmatched
            st.session_state.source_pages_count = len(source_urls)
        
        if st.session_state.rc_df is not None:
            source_urls = st.session_state.rc_df['Source'].unique().tolist()
            matched, unmatched = match_post_ids_to_sources(source_urls)
            st.session_state.post_id_matched_count = matched
            st.session_state.post_id_unmatched_urls = unmatched
            st.session_state.source_pages_count = len(source_urls)
        
        return True
    else:
        st.error("Could not find Post IDs in this file. Make sure it has an 'Address' column and a 'post_id' (or 'post_id 1') column.")
        return False


def process_redirect_chains_upload(uploaded_file) -> bool:
    """Process an uploaded redirect chains CSV. Returns True if processing succeeded."""
    df = parse_redirect_chains_csv(uploaded_file)
    if df is not None and len(df) > 0:
        domain = detect_domain(df['Source'].tolist())
        redirects, sitewide, loops = group_redirect_chains(df, domain)
        
        # Initialize decisions for each redirect
        decisions = {}
        for key in redirects:
            decisions[key] = {
                'approved_action': '',
                'approved_fix': '',
            }
        
        # Store in session state
        st.session_state.rc_df = df
        st.session_state.rc_domain = domain
        st.session_state.rc_redirects = redirects
        st.session_state.rc_decisions = decisions
        st.session_state.rc_sitewide = sitewide
        st.session_state.rc_loops = loops
        st.session_state.task_type = 'redirect_chains'
        st.session_state.current_task = 'redirect_chains'
        
        # Count unique source pages
        source_urls = df['Source'].unique().tolist()
        source_pages_count = len(source_urls)
        st.session_state.source_pages_count = source_pages_count
        
        # Check if CSV itself has post_id column
        csv_has_post_ids = df['post_id'].notna().any()
        if csv_has_post_ids:
            for _, row in df.iterrows():
                if pd.notna(row.get('post_id')):
                    st.session_state.post_id_cache[row['Source']] = int(row['post_id'])
            st.session_state.has_post_ids = True
        
        # Match Post IDs if we have them from separate file
        if st.session_state.post_id_file_uploaded:
            matched, unmatched = match_post_ids_to_sources(source_urls)
            st.session_state.post_id_matched_count = matched
            st.session_state.post_id_unmatched_urls = unmatched
        
        # Update mode based on Post ID availability
        st.session_state.selected_mode = 'full' if st.session_state.has_post_ids else 'quick_start'
        
        # Track upload
        total_redirects = len(redirects)
        temp_count = sum(1 for r in redirects.values() if r['is_temp_redirect'])
        track_event("csv_upload", {
            "type": "redirect_chains",
            "unique_redirects": total_redirects,
            "total_references": len(df),
            "source_pages": source_pages_count,
            "sitewide_count": len(sitewide),
            "loop_count": len(loops),
            "temp_redirect_count": temp_count,
            "has_post_ids": st.session_state.has_post_ids
        })
        
        return True
    return False


def process_broken_links_upload(uploaded_file) -> bool:
    """Process an uploaded broken links CSV. Returns True if processing succeeded."""
    df = parse_csv(uploaded_file)
    if df is not None and len(df) > 0:
        domain = detect_domain(df['Source'].tolist())
        broken_urls = group_by_broken_url(df, domain)
        
        decisions = {}
        for url in broken_urls:
            decisions[url] = {
                'manual_fix': '',
                'ai_suggestion': '',
                'ai_action': '',
                'ai_notes': '',
                'approved_fix': '',
                'approved_action': '',
            }
        
        st.session_state.df = df
        st.session_state.domain = domain
        st.session_state.broken_urls = broken_urls
        st.session_state.decisions = decisions
        st.session_state.task_type = 'broken_links'
        st.session_state.current_task = 'broken_links'
        
        # Count unique source pages
        source_urls = df['Source'].unique().tolist()
        source_pages_count = len(source_urls)
        st.session_state.source_pages_count = source_pages_count
        
        # Check if CSV itself has post_id column
        csv_has_post_ids = df['post_id'].notna().any()
        if csv_has_post_ids:
            for _, row in df.iterrows():
                if pd.notna(row.get('post_id')):
                    st.session_state.post_id_cache[row['Source']] = int(row['post_id'])
            st.session_state.has_post_ids = True
        
        # Match Post IDs if we have them from separate file
        if st.session_state.post_id_file_uploaded:
            matched, unmatched = match_post_ids_to_sources(source_urls)
            st.session_state.post_id_matched_count = matched
            st.session_state.post_id_unmatched_urls = unmatched
        
        # Update mode based on Post ID availability
        st.session_state.selected_mode = 'full' if st.session_state.has_post_ids else 'quick_start'
        
        # Track CSV upload
        track_event("csv_upload", {
            "type": "broken_links",
            "unique_broken_urls": len(broken_urls),
            "total_references": len(df),
            "source_pages": source_pages_count,
            "has_post_ids": st.session_state.has_post_ids
        })
        
        return True
    return False


def process_image_alt_text_upload(uploaded_file) -> bool:
    """Process an uploaded Image Alt Text CSV (All Image Inlinks from Screaming Frog). Returns True if processing succeeded."""
    df = parse_image_alt_text_csv(uploaded_file)
    if df is not None and len(df) > 0:
        domain = detect_domain(df['Source'].tolist())
        images, excluded_count = group_images_for_alt_text(df, domain)
        
        if not images:
            st.warning("No images found that need alt text fixes after filtering. All images either have good alt text or were excluded (logos, icons, non-content pages, etc.)")
            return False
        
        decisions = {}
        for img_url in images:
            decisions[img_url] = {
                'manual_fix': '',
                'ai_suggestion': '',
                'ai_notes': '',
                'approved_fix': '',
                'approved_action': '',  # 'replace' or 'ignore'
            }
        
        st.session_state.iat_df = df
        st.session_state.iat_domain = domain
        st.session_state.iat_images = images
        st.session_state.iat_decisions = decisions
        st.session_state.iat_excluded_count = excluded_count
        st.session_state.task_type = 'image_alt_text'
        st.session_state.current_task = 'image_alt_text'
        
        # Count unique source pages
        all_sources = set()
        for img_data in images.values():
            all_sources.update(img_data['sources'])
        source_pages_count = len(all_sources)
        st.session_state.source_pages_count = source_pages_count  # Store for mode selector
        
        # Check if CSV itself has post_id column
        csv_has_post_ids = df['post_id'].notna().any() if 'post_id' in df.columns else False
        if csv_has_post_ids:
            for _, row in df.iterrows():
                if pd.notna(row.get('post_id')):
                    st.session_state.post_id_cache[row['Source']] = int(row['post_id'])
            st.session_state.has_post_ids = True
        
        # Match Post IDs if we have them from separate file
        if st.session_state.post_id_file_uploaded:
            matched, unmatched = match_post_ids_to_sources(list(all_sources))
            st.session_state.post_id_matched_count = matched
            st.session_state.post_id_unmatched_urls = unmatched
        
        # Update mode based on Post ID availability
        st.session_state.selected_mode = 'full' if st.session_state.has_post_ids else 'quick_start'
        
        # Track CSV upload
        track_event("csv_upload", {
            "type": "image_alt_text",
            "unique_images": len(images),
            "total_references": len(df),
            "excluded_count": excluded_count,
            "source_pages": source_pages_count,
            "has_post_ids": st.session_state.has_post_ids
        })
        
        return True
    return False


def render_post_id_match_status(total_pages: int):
    """Render the Post ID matching status after report upload"""
    matched = st.session_state.post_id_matched_count
    unmatched = st.session_state.post_id_unmatched_urls
    unmatched_summary = get_unmatched_summary(unmatched)
    
    if total_pages > 0:
        match_pct = (matched / total_pages) * 100
    else:
        match_pct = 0
    
    # Build the unmatched section if needed
    unmatched_html = ''
    if unmatched_summary['total'] > 0:
        unmatched_html = f'''
        <div style="font-size: 0.9rem; color: #15803d;">
            ℹ️ {unmatched_summary['total']} URLs unmatched — this is expected:
            <ul style="margin: 0.25rem 0 0 1.5rem; padding: 0;">
                <li>Category, tag, author, and archive pages don't have Post IDs</li>
                <li>These are dynamically generated by WordPress</li>
                <li>Fix the individual posts and these pages update automatically</li>
            </ul>
        </div>
        '''
    
    st.markdown(f'''
    <div style="background: #f0fdf4; padding: 1rem; border-radius: 8px; border: 1px solid #bbf7d0; margin: 0.5rem 0;">
        <div style="font-weight: 600; color: #166534; margin-bottom: 0.5rem;">
            ✅ {matched} of {total_pages} URLs matched with Post IDs ({match_pct:.0f}%)
        </div>
        {unmatched_html}
    </div>
    ''', unsafe_allow_html=True)


def render_post_id_extraction_guide():
    """Render detailed Screaming Frog Post ID extraction instructions"""
    st.markdown("""
    ### Why Post IDs Matter
    
    When you see a URL like `/how-to-start-a-food-truck/`, that's the human-friendly version. 
    But WordPress uses **numeric Post IDs** internally (like `6125`). To edit content via the API, 
    we need this number.
    
    **Good news:** Your site already contains Post IDs in the HTML. We just need Screaming Frog to extract them.
    
    **⚠️ Requires a licensed version of Screaming Frog** (~$259/year) for Custom Extraction. 
    [Get it here](https://www.screamingfrog.co.uk/seo-spider/) — incredible value for professional SEO.
    
    ---
    
    ### Quick Setup (5 minutes, one-time)
    
    **Step 1:** In Screaming Frog, go to **Configuration → Custom → Extraction**
    
    **Step 2:** Click **Add** and configure:
    - **Name:** `post_id`
    - **Type:** Change to **Regex**
    
    **Step 3:** Add the Regex pattern (see options below)
    
    **Step 4:** Save as default: **Configuration → Profiles → Save Current Configuration as Default**
    
    ---
    
    ### Which Regex Pattern Do I Use?
    
    Check your site's HTML source (View Page Source) and look for one of these patterns:
    """)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **Option A: Shortlink** *(most common)*
        
        Look for: `<link rel="shortlink" href="...?p=6125">`
        
        ```
        <link[^>]+rel=['"]shortlink['"][^>]+href=['"][^'"]*\\?p=(\\d+)
        ```
        
        **Option B: Body Class**
        
        Look for: `<body class="...postid-6125...">`
        
        ```
        class=['"][^'"]*(?:postid|page-id)-(\\d+)
        ```
        """)
    
    with col2:
        st.markdown("""
        **Option C: Article ID**
        
        Look for: `<article id="post-6125">`
        
        ```
        <article[^>]+id=['"]post-(\\d+)
        ```
        
        **Option D: REST API Link**
        
        Look for: `wp-json/wp/v2/posts/6125`
        
        ```
        wp-json/wp/v2/posts/(\\d+)
        ```
        """)
    
    st.markdown("---")
    
    with st.expander("🤖 Can't find your pattern? Use this AI prompt"):
        st.markdown("""
        Copy this prompt and paste it into ChatGPT, Claude, or any AI assistant along with 
        50-100 lines of your page's HTML source:
        """)
        
        st.code("""I need to extract WordPress Post IDs from my website's HTML using Screaming Frog's Custom Extraction feature with Regex.

Here is a sample of my page's HTML source code:

[PASTE YOUR HTML HERE]

Please analyze this HTML and:
1. Identify where the WordPress Post ID is stored
2. Provide the exact Regex pattern for Screaming Frog""", language=None)
    
    st.markdown("""
    ---
    
    📖 **[View Full Setup Guide](https://github.com/backofnapkin/screaming-fixes/blob/main/POST_ID_SETUP.md)** — includes troubleshooting, screenshots, and video walkthrough.
    """)


def render_metrics():
    """Render summary metrics"""
    broken_urls = st.session_state.broken_urls
    decisions = st.session_state.decisions
    
    total = len(broken_urls)
    internal = sum(1 for info in broken_urls.values() if info['is_internal'])
    external = total - internal
    approved = sum(1 for d in decisions.values() if d['approved_action'])
    pending = total - approved
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Broken", total)
    with col2:
        st.metric("Internal", internal)
    with col3:
        st.metric("External", external)
    with col4:
        st.metric("✅ Approved", approved)


def render_spreadsheet():
    """Render the main spreadsheet view - compact inline editing"""
    broken_urls = st.session_state.broken_urls
    decisions = st.session_state.decisions
    domain = st.session_state.domain
    
    # Count pending for bulk actions
    pending_urls = [url for url, d in decisions.items() if not d['approved_action']]
    has_ai_key = bool(st.session_state.anthropic_key) or (AGENT_MODE_API_KEY and st.session_state.ai_suggestions_remaining > 0)
    
    # Top action bar
    st.markdown("---")
    action_cols = st.columns([2, 2, 2])
    
    with action_cols[0]:
        if has_ai_key and len(pending_urls) > 0:
            if st.button(f"🤖 Bulk AI Analyze ({len(pending_urls)})", type="primary", use_container_width=True, help="Get AI suggestions for all pending URLs"):
                st.session_state.show_bulk_ai_modal = True
                st.rerun()
        else:
            st.button(f"🤖 Bulk AI Analyze", use_container_width=True, disabled=True, help="Add API key to enable")
    
    with action_cols[1]:
        # Count URLs with AI suggestions ready to approve
        ai_ready = [url for url, d in decisions.items() if d['ai_action'] and not d['approved_action']]
        if ai_ready:
            if st.button(f"✅ Approve All AI ({len(ai_ready)})", use_container_width=True, help="Approve all pending AI suggestions"):
                for url in ai_ready:
                    decisions[url]['approved_action'] = decisions[url]['ai_action']
                    decisions[url]['approved_fix'] = decisions[url]['ai_suggestion']
                st.toast(f"✅ Approved {len(ai_ready)} AI suggestions", icon="✅")
                time.sleep(0.3)
                st.rerun()
        else:
            st.button("✅ Approve All AI", use_container_width=True, disabled=True, help="No AI suggestions to approve")
    
    with action_cols[2]:
        approved_count = sum(1 for d in decisions.values() if d['approved_action'])
        if st.button(f"📥 Export ({approved_count})", use_container_width=True, disabled=approved_count == 0):
            st.session_state.scroll_to_export = True
    
    # Handle bulk AI modal
    if st.session_state.get('show_bulk_ai_modal'):
        render_bulk_ai_modal(pending_urls, broken_urls, domain)
        return
    
    st.markdown("---")
    
    # Compact filters row - consolidated dropdowns
    filter_cols = st.columns([1.5, 1.5, 1.2, 1.2, 1.5])
    
    with filter_cols[0]:
        # Link type filter (All/Internal/External)
        link_type_options = ["All Links", "Internal Only", "External Only"]
        current_link_type = "All Links"
        if st.session_state.get('filter_internal', True) and not st.session_state.get('filter_external', True):
            current_link_type = "Internal Only"
        elif not st.session_state.get('filter_internal', True) and st.session_state.get('filter_external', True):
            current_link_type = "External Only"
        
        selected_link_type = st.selectbox(
            "Link Type",
            link_type_options,
            index=link_type_options.index(current_link_type),
            key="link_type_filter",
            label_visibility="collapsed"
        )
        # Update session state based on selection
        if selected_link_type == "All Links":
            st.session_state.filter_internal = True
            st.session_state.filter_external = True
        elif selected_link_type == "Internal Only":
            st.session_state.filter_internal = True
            st.session_state.filter_external = False
        else:  # External Only
            st.session_state.filter_internal = False
            st.session_state.filter_external = True
    
    with filter_cols[1]:
        # Status code filter
        status_codes = sorted(set(info['status_code'] for info in broken_urls.values()))
        status_options = ["All Status Codes"] + [str(code) for code in status_codes]
        selected_status = st.selectbox(
            "Status",
            status_options,
            index=0,
            key="status_filter",
            label_visibility="collapsed"
        )
        st.session_state.filter_status = None if selected_status == "All Status Codes" else int(selected_status)
    
    with filter_cols[2]:
        st.session_state.show_pending = st.checkbox("Pending", value=st.session_state.get('show_pending', True), key="f_pend")
    
    with filter_cols[3]:
        st.session_state.show_approved = st.checkbox("Approved", value=st.session_state.get('show_approved', True), key="f_appr")
    
    with filter_cols[4]:
        sort_option = st.selectbox(
            "Sort",
            ["Impact ↓", "Status Code", "Internal First", "External First"],
            index=0,
            key="sort_select",
            label_visibility="collapsed"
        )
    
    # Apply filters
    filtered_urls = []
    for url, info in broken_urls.items():
        if info['is_internal'] and not st.session_state.filter_internal:
            continue
        if not info['is_internal'] and not st.session_state.filter_external:
            continue
        # Status code filter
        if st.session_state.get('filter_status') and info['status_code'] != st.session_state.filter_status:
            continue
        has_approval = bool(decisions[url]['approved_action'])
        if has_approval and not st.session_state.show_approved:
            continue
        if not has_approval and not st.session_state.show_pending:
            continue
        filtered_urls.append(url)
    
    if not filtered_urls:
        st.info("No URLs match your filters")
        return
    
    # Apply sorting
    sort_map = {"Impact ↓": "impact", "Status Code": "status_code", "Internal First": "internal_first", "External First": "external_first"}
    sort_by = sort_map.get(sort_option, "impact")
    
    if sort_by == "impact":
        filtered_urls.sort(key=lambda u: broken_urls[u]['count'], reverse=True)
    elif sort_by == "status_code":
        filtered_urls.sort(key=lambda u: broken_urls[u]['status_code'])
    elif sort_by == "internal_first":
        filtered_urls.sort(key=lambda u: (0 if broken_urls[u]['is_internal'] else 1, -broken_urls[u]['count']))
    elif sort_by == "external_first":
        filtered_urls.sort(key=lambda u: (1 if broken_urls[u]['is_internal'] else 0, -broken_urls[u]['count']))
    
    # Pagination
    total = len(filtered_urls)
    per_page = 15  # More rows visible
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(st.session_state.page, total_pages - 1)
    
    start = page * per_page
    end = min(start + per_page, total)
    page_urls = filtered_urls[start:end]
    
    # Build descriptive header based on active filters
    # Link type description
    if st.session_state.filter_internal and st.session_state.filter_external:
        link_type_desc = ""
    elif st.session_state.filter_internal:
        link_type_desc = "internal "
    else:
        link_type_desc = "external "
    
    # Status code description
    status_filter = st.session_state.get('filter_status')
    if status_filter:
        status_desc = f" with {status_filter} errors"
    else:
        status_desc = ""
    
    # Approval status description
    if st.session_state.show_pending and st.session_state.show_approved:
        approval_desc = ""
    elif st.session_state.show_pending:
        approval_desc = " (pending only)"
    else:
        approval_desc = " (approved only)"
    
    st.markdown(f"**Showing {start+1}-{end} of {total}** {link_type_desc}broken links{status_desc}{approval_desc}")
    
    # Column headers - styled like a spreadsheet with teal branding
    st.markdown("""
    <div style="display: flex; background: linear-gradient(135deg, #f0fdfa 0%, #ccfbf1 100%); border: 1px solid #99f6e4; border-radius: 8px; padding: 0.5rem 0.75rem; margin: 0.5rem 0;">
        <div style="flex: 4; font-size: 0.75rem; color: #0d9488; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em;">Broken Link</div>
        <div style="flex: 4; font-size: 0.75rem; color: #0d9488; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em;">Fix</div>
        <div style="flex: 1; font-size: 0.75rem; color: #0d9488; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; text-align: center;">Approve</div>
        <div style="flex: 0.8; font-size: 0.75rem; color: #0d9488; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; text-align: center;">Edit</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Render compact rows
    for url in page_urls:
        render_compact_url_row(url, broken_urls[url], decisions[url], domain)
    
    # Pagination
    if total_pages > 1:
        pcol1, pcol2, pcol3 = st.columns([1, 2, 1])
        with pcol1:
            if st.button("← Prev", disabled=page == 0, key="prev"):
                st.session_state.page = page - 1
                st.rerun()
        with pcol2:
            st.markdown(f"<div style='text-align:center; padding-top: 0.5rem;'>Page {page+1} / {total_pages}</div>", unsafe_allow_html=True)
        with pcol3:
            if st.button("Next →", disabled=page >= total_pages - 1, key="next"):
                st.session_state.page = page + 1
                st.rerun()


def get_batch_ai_suggestions(urls_batch: List[Dict], domain: str, api_key: str) -> List[Dict]:
    """Get AI suggestions for a batch of URLs (up to 10 at a time)"""
    
    if not ANTHROPIC_AVAILABLE:
        return [{'url': item['url'], 'action': 'remove', 'replacement': None, 'notes': 'Anthropic library not installed.'} for item in urls_batch]
    
    try:
        client = Anthropic(api_key=api_key)
        
        # Build batch prompt
        urls_list = []
        for i, item in enumerate(urls_batch, 1):
            anchors = ', '.join(f'"{a}"' for a in item['anchors'][:3]) if item['anchors'] else 'none'
            link_type = "internal" if item['is_internal'] else "external"
            urls_list.append(f"{i}. {item['url']} (status: {item['status_code']}, type: {link_type}, anchors: {anchors}, affects: {item['count']} pages)")
        
        urls_text = '\n'.join(urls_list)
        
        prompt = f"""You are helping fix broken links on {domain}.

Here are {len(urls_batch)} broken URLs to analyze:

{urls_text}

For each URL, determine if we should:
- REMOVE: Delete the link but keep anchor text (use when no good replacement exists)
- REPLACE: Replace with a working URL (only if you find a real replacement)

Use web search to check if content has moved to new URLs.

Respond with a JSON array, one object per URL in order:
[
  {{"index": 1, "action": "remove" or "replace", "url": "replacement URL or null", "notes": "brief 1-sentence explanation"}},
  {{"index": 2, "action": "...", "url": "...", "notes": "..."}},
  ...
]

Only output the JSON array, nothing else."""

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            tools=[{
                "type": "web_search_20250305",
                "name": "web_search",
                "max_uses": 10
            }],
            messages=[{"role": "user", "content": prompt}]
        )
        
        result_text = ""
        for block in response.content:
            if hasattr(block, 'text'):
                result_text += block.text
        
        # Parse JSON array response
        try:
            json_match = re.search(r'\[[\s\S]*\]', result_text)
            if json_match:
                results = json.loads(json_match.group())
                # Map results back to URLs
                output = []
                for i, item in enumerate(urls_batch):
                    if i < len(results):
                        r = results[i]
                        output.append({
                            'url': item['url'],
                            'action': r.get('action', 'remove'),
                            'replacement': r.get('url'),
                            'notes': r.get('notes', 'No explanation provided.')
                        })
                    else:
                        output.append({
                            'url': item['url'],
                            'action': 'remove',
                            'replacement': None,
                            'notes': 'Could not analyze this URL.'
                        })
                return output
        except json.JSONDecodeError:
            pass
        
        # Fallback if parsing fails
        return [{'url': item['url'], 'action': 'remove', 'replacement': None, 'notes': 'Could not parse AI response.'} for item in urls_batch]
        
    except Exception as e:
        error_msg = str(e)[:100]
        # Check for rate limit
        if 'rate' in error_msg.lower() or '429' in error_msg:
            raise Exception("RATE_LIMIT")
        return [{'url': item['url'], 'action': 'remove', 'replacement': None, 'notes': f'Error: {error_msg}'} for item in urls_batch]


def render_bulk_ai_modal(pending_urls: List[str], broken_urls: Dict, domain: str):
    """Render bulk AI analysis modal with batch processing"""
    
    # Get already analyzed URLs (those with AI suggestions)
    already_analyzed = st.session_state.get('bulk_ai_analyzed_urls', set())
    
    # Filter to only unanalyzed pending URLs
    unanalyzed_urls = [url for url in pending_urls if url not in already_analyzed]
    
    # Sort by impact (page count)
    unanalyzed_urls.sort(key=lambda u: broken_urls[u]['count'], reverse=True)
    
    total_pending = len(pending_urls)
    total_unanalyzed = len(unanalyzed_urls)
    
    # Calculate page impact for different tiers
    def get_impact(url_list):
        return sum(broken_urls[u]['count'] for u in url_list)
    
    top_25 = unanalyzed_urls[:25]
    top_50 = unanalyzed_urls[:50]
    top_100 = unanalyzed_urls[:100]
    
    impact_25 = get_impact(top_25)
    impact_50 = get_impact(top_50)
    impact_100 = get_impact(top_100)
    total_impact = get_impact(unanalyzed_urls)
    
    # Check if we're in progress
    if st.session_state.get('bulk_ai_running'):
        render_bulk_ai_progress(unanalyzed_urls, broken_urls, domain)
        return
    
    # Check if we just completed
    if st.session_state.get('bulk_ai_just_completed'):
        render_bulk_ai_completion(total_pending, total_unanalyzed, broken_urls)
        return
    
    # Launch modal
    st.markdown("### 🤖 Bulk AI Analysis")
    
    if total_unanalyzed == 0:
        st.success("✅ All pending URLs have been analyzed!")
        if st.button("Close", use_container_width=True):
            st.session_state.show_bulk_ai_modal = False
            st.rerun()
        return
    
    st.markdown(f"**{total_unanalyzed}** broken URLs ready to analyze" + (f" ({len(already_analyzed)} already analyzed)" if already_analyzed else ""))
    
    st.markdown("**Analyze by impact:**")
    
    # Tier selection
    tier_options = []
    if len(top_25) > 0:
        tier_options.append(f"Top 25 (affects {impact_25:,} pages)")
    if len(top_50) > 25:
        tier_options.append(f"Top 50 (affects {impact_50:,} pages)")
    if len(top_100) > 50:
        tier_options.append(f"Top 100 (affects {impact_100:,} pages)")
    
    # Add "All" option (disabled if > 100)
    if total_unanalyzed > 100:
        tier_options.append(f"All {total_unanalyzed} — Max 100 per batch")
    elif total_unanalyzed > 0 and f"Top {total_unanalyzed}" not in str(tier_options):
        tier_options.append(f"All {total_unanalyzed} (affects {total_impact:,} pages)")
    
    # Default to highest available tier up to 100
    default_idx = min(len(tier_options) - 1, 2) if len(tier_options) > 2 else len(tier_options) - 1
    
    selected_tier = st.radio(
        "Select batch size",
        tier_options,
        index=default_idx,
        label_visibility="collapsed"
    )
    
    # Determine actual count
    if "Top 25" in selected_tier:
        analyze_count = 25
    elif "Top 50" in selected_tier:
        analyze_count = 50
    elif "Top 100" in selected_tier or "Max 100" in selected_tier:
        analyze_count = min(100, total_unanalyzed)
    else:
        analyze_count = min(total_unanalyzed, 100)
    
    # Time estimate (batch of 10 takes ~8 seconds)
    batches = (analyze_count + 9) // 10
    est_seconds = batches * 8
    est_time = f"{est_seconds // 60}m {est_seconds % 60}s" if est_seconds >= 60 else f"{est_seconds}s"
    
    st.markdown(f"⏱️ Estimated time: **~{est_time}**")
    
    # Tip about batching
    if total_unanalyzed > 100:
        st.info("💡 **Need to analyze more than 100?** Run in batches! After this batch completes, click \"Analyze Next Batch\" to continue. Your previous results are saved.")
    
    st.markdown("")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("🚀 Start Analysis", type="primary", use_container_width=True):
            st.session_state.bulk_ai_running = True
            st.session_state.bulk_ai_progress = 0
            st.session_state.bulk_ai_total = analyze_count
            st.session_state.bulk_ai_urls_to_process = unanalyzed_urls[:analyze_count]
            st.session_state.bulk_ai_results_summary = {'replace': 0, 'remove': 0, 'error': 0}
            st.session_state.bulk_ai_start_time = time.time()
            if 'bulk_ai_analyzed_urls' not in st.session_state:
                st.session_state.bulk_ai_analyzed_urls = set()
            st.rerun()
    with col2:
        if st.button("Cancel", use_container_width=True):
            st.session_state.show_bulk_ai_modal = False
            st.rerun()


def render_bulk_ai_progress(unanalyzed_urls: List[str], broken_urls: Dict, domain: str):
    """Render the progress view during bulk AI analysis with detailed feedback"""
    
    progress = st.session_state.get('bulk_ai_progress', 0)
    total = st.session_state.get('bulk_ai_total', 100)
    urls_to_process = st.session_state.get('bulk_ai_urls_to_process', [])
    results_summary = st.session_state.get('bulk_ai_results_summary', {'replace': 0, 'remove': 0, 'error': 0})
    start_time = st.session_state.get('bulk_ai_start_time', time.time())
    recent_results = st.session_state.get('bulk_ai_recent_results', [])
    
    # Check for paused state (rate limit)
    paused_until = st.session_state.get('bulk_ai_paused_until', 0)
    pause_reason = st.session_state.get('bulk_ai_pause_reason', '')
    
    # Check for error state
    error_state = st.session_state.get('bulk_ai_error_state', None)
    
    # Handle rate limit pause with countdown
    if paused_until and time.time() < paused_until:
        remaining_wait = int(paused_until - time.time())
        
        st.markdown("### ⏸️ Rate Limit Reached")
        st.markdown("""
        Your AI provider limits how many requests can be made per minute. 
        This is normal for large batches.
        """)
        
        # Countdown progress bar
        total_wait = 60
        wait_progress = (total_wait - remaining_wait) / total_wait
        st.progress(wait_progress)
        st.markdown(f"⏱️ Auto-resuming in: **{remaining_wait} seconds**")
        
        st.markdown(f"Progress saved: **{progress} of {total}** complete")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("▶️ Resume Now", use_container_width=True):
                st.session_state.bulk_ai_paused_until = 0
                st.session_state.bulk_ai_pause_reason = ''
                st.rerun()
        with col2:
            if st.button("⏹️ Stop & Keep Results", use_container_width=True):
                st.session_state.bulk_ai_running = False
                st.session_state.bulk_ai_paused_until = 0
                st.session_state.bulk_ai_just_completed = True
                st.rerun()
        
        # Auto-refresh countdown
        time.sleep(1)
        st.rerun()
        return
    
    # Clear pause state if time has passed
    if paused_until and time.time() >= paused_until:
        st.session_state.bulk_ai_paused_until = 0
        st.session_state.bulk_ai_pause_reason = ''
    
    # Handle error states
    if error_state:
        error_type = error_state.get('type', 'unknown')
        error_msg = error_state.get('message', 'An unknown error occurred')
        error_batch = error_state.get('batch', 0)
        
        if error_type == 'timeout':
            st.markdown("### ⚠️ Batch Timed Out")
            st.markdown(f"""
            The AI took too long to respond (>45 seconds). This can happen 
            when web searches are slow or the URLs are complex.
            
            **Batch {error_batch}** will be skipped and those URLs marked for manual review.
            """)
        elif error_type == 'connection':
            st.markdown("### ❌ Connection Error")
            st.markdown(f"""
            Could not reach the AI service. This might be:
            - Temporary network issue
            - AI service is down
            - Invalid API key
            
            Error: `{error_msg}`
            """)
        else:
            st.markdown("### ❌ Error Occurred")
            st.markdown(f"Error: `{error_msg}`")
        
        st.markdown(f"Progress saved: **{progress} of {total}** complete")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🔄 Retry Batch", type="primary", use_container_width=True):
                st.session_state.bulk_ai_error_state = None
                st.rerun()
        with col2:
            if st.button("⏭️ Skip & Continue", use_container_width=True):
                # Skip current batch
                batch_size = min(10, total - progress)
                st.session_state.bulk_ai_progress = progress + batch_size
                st.session_state.bulk_ai_error_state = None
                # Mark skipped URLs
                for i in range(batch_size):
                    if progress + i < len(urls_to_process):
                        url = urls_to_process[progress + i]
                        st.session_state.decisions[url]['ai_action'] = ''
                        st.session_state.decisions[url]['ai_notes'] = 'Skipped due to error - review manually'
                        results_summary['error'] += 1
                st.session_state.bulk_ai_results_summary = results_summary
                st.rerun()
        
        if st.button("⏹️ Stop & Keep Results", use_container_width=True):
            st.session_state.bulk_ai_running = False
            st.session_state.bulk_ai_error_state = None
            st.session_state.bulk_ai_just_completed = True
            st.rerun()
        return
    
    # Normal progress view
    st.markdown("### 🤖 Analyzing Broken Links...")
    
    # Progress bar
    progress_pct = progress / total if total > 0 else 0
    st.progress(progress_pct)
    st.markdown(f"**{progress} of {total}** URLs analyzed")
    
    # Time tracking
    elapsed = time.time() - start_time
    if progress > 0:
        rate = elapsed / progress
        remaining = (total - progress) * rate
        elapsed_str = f"{int(elapsed // 60)}:{int(elapsed % 60):02d}"
        remaining_str = f"{int(remaining // 60)}:{int(remaining % 60):02d}"
        st.markdown(f"⏱️ Elapsed: **{elapsed_str}** | Remaining: **~{remaining_str}**")
    
    # Batch info
    current_batch = (progress // 10) + 1
    total_batches = (total + 9) // 10
    batch_start = progress + 1
    batch_end = min(progress + 10, total)
    
    st.markdown(f"📦 **Processing batch {current_batch} of {total_batches}** (URLs {batch_start}-{batch_end})")
    
    # AI "thinking" animation
    thinking_states = [
        "🔍 Searching web for replacement URLs...",
        "🌐 Checking if content has moved...",
        "📊 Analyzing anchor text context...",
        "🔗 Validating potential replacements...",
        "💭 Determining best action..."
    ]
    thinking_idx = int(time.time() * 0.5) % len(thinking_states)
    
    st.markdown(f"""
    <div style="background: #f0fdfa; border: 1px solid #99f6e4; border-radius: 8px; padding: 0.75rem 1rem; margin: 0.5rem 0;">
        <span style="color: #0d9488;">{thinking_states[thinking_idx]}</span>
    </div>
    """, unsafe_allow_html=True)
    
    # Recent results feed
    if recent_results:
        st.markdown("**Recent results:**")
        for result in recent_results[-5:]:  # Show last 5
            url_short = result['url'].replace('https://', '').replace('http://', '')
            if len(url_short) > 40:
                url_short = url_short[:37] + "..."
            
            if result['action'] == 'replace':
                replacement_short = result.get('replacement', '')
                if len(replacement_short) > 30:
                    replacement_short = replacement_short[:27] + "..."
                st.markdown(f"- `{url_short}` → **Replace** → `{replacement_short}`")
            else:
                reason = result.get('notes', 'no replacement found')
                if len(reason) > 40:
                    reason = reason[:37] + "..."
                st.markdown(f"- `{url_short}` → **Remove** ({reason})")
    
    # Results tally
    st.markdown("---")
    st.markdown("**Totals:**")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Replace", results_summary['replace'])
    with col2:
        st.metric("Remove", results_summary['remove'])
    with col3:
        st.metric("Undetermined", results_summary['error'])
    
    # Stop button
    if st.button("⏹️ Stop & Keep Results", use_container_width=True):
        st.session_state.bulk_ai_running = False
        st.session_state.bulk_ai_just_completed = True
        st.rerun()
    
    # Process next batch
    if progress < total:
        batch_urls = urls_to_process[progress:min(progress + 10, total)]
        
        if batch_urls:
            # Prepare batch data
            batch_data = []
            for url in batch_urls:
                info = broken_urls[url]
                batch_data.append({
                    'url': url,
                    'status_code': info['status_code'],
                    'is_internal': info['is_internal'],
                    'anchors': info['anchors'],
                    'count': info['count']
                })
            
            # Get API key
            api_key = st.session_state.anthropic_key or AGENT_MODE_API_KEY
            
            if api_key:
                try:
                    # Process batch with timeout
                    import concurrent.futures
                    
                    with concurrent.futures.ThreadPoolExecutor() as executor:
                        future = executor.submit(get_batch_ai_suggestions, batch_data, domain, api_key)
                        try:
                            results = future.result(timeout=45)  # 45 second timeout
                        except concurrent.futures.TimeoutError:
                            st.session_state.bulk_ai_error_state = {
                                'type': 'timeout',
                                'message': 'Request took longer than 45 seconds',
                                'batch': current_batch
                            }
                            st.rerun()
                            return
                    
                    # Apply results
                    new_recent = []
                    for result in results:
                        url = result['url']
                        st.session_state.decisions[url]['ai_action'] = result['action']
                        st.session_state.decisions[url]['ai_suggestion'] = result['replacement'] or ''
                        st.session_state.decisions[url]['ai_notes'] = result['notes']
                        st.session_state.bulk_ai_analyzed_urls.add(url)
                        
                        # Track for recent results display
                        new_recent.append({
                            'url': url,
                            'action': result['action'],
                            'replacement': result['replacement'],
                            'notes': result['notes']
                        })
                        
                        # Update summary
                        if result['action'] == 'replace' and result['replacement']:
                            results_summary['replace'] += 1
                        elif result['action'] == 'remove':
                            results_summary['remove'] += 1
                        else:
                            results_summary['error'] += 1
                    
                    # Update recent results
                    recent_results.extend(new_recent)
                    st.session_state.bulk_ai_recent_results = recent_results[-10:]  # Keep last 10
                    
                    st.session_state.bulk_ai_results_summary = results_summary
                    st.session_state.bulk_ai_progress = progress + len(batch_urls)
                    
                    # Decrement free suggestions if using agent key
                    if not st.session_state.anthropic_key and AGENT_MODE_API_KEY:
                        st.session_state.ai_suggestions_remaining = max(0, st.session_state.ai_suggestions_remaining - len(batch_urls))
                    
                    st.rerun()
                    
                except Exception as e:
                    error_str = str(e)
                    
                    if "RATE_LIMIT" in error_str or "429" in error_str or "rate" in error_str.lower():
                        # Set pause state with countdown
                        st.session_state.bulk_ai_paused_until = time.time() + 60
                        st.session_state.bulk_ai_pause_reason = 'rate_limit'
                        st.rerun()
                    elif "timeout" in error_str.lower() or "timed out" in error_str.lower():
                        st.session_state.bulk_ai_error_state = {
                            'type': 'timeout',
                            'message': error_str[:100],
                            'batch': current_batch
                        }
                        st.rerun()
                    elif "connection" in error_str.lower() or "network" in error_str.lower():
                        st.session_state.bulk_ai_error_state = {
                            'type': 'connection',
                            'message': error_str[:100],
                            'batch': current_batch
                        }
                        st.rerun()
                    else:
                        st.session_state.bulk_ai_error_state = {
                            'type': 'unknown',
                            'message': error_str[:150],
                            'batch': current_batch
                        }
                        st.rerun()
    else:
        # Done!
        st.session_state.bulk_ai_running = False
        st.session_state.bulk_ai_just_completed = True
        st.rerun()


def render_bulk_ai_completion(total_pending: int, total_analyzed: int, broken_urls: Dict):
    """Render the completion view after bulk AI analysis"""
    
    results_summary = st.session_state.get('bulk_ai_results_summary', {'replace': 0, 'remove': 0, 'error': 0})
    analyzed_urls = st.session_state.get('bulk_ai_analyzed_urls', set())
    
    st.markdown("### ✅ Bulk Analysis Complete")
    
    analyzed_count = results_summary['replace'] + results_summary['remove'] + results_summary['error']
    remaining = total_pending - len(analyzed_urls)
    
    st.markdown(f"Analyzed **{analyzed_count}** of **{total_pending}** broken URLs")
    
    # Results breakdown
    st.markdown("**Results:**")
    st.markdown(f"- **{results_summary['replace']}** → Replace with new URL")
    st.markdown(f"- **{results_summary['remove']}** → Remove link")
    if results_summary['error'] > 0:
        st.markdown(f"- **{results_summary['error']}** → Could not determine (review manually)")
    
    # Impact summary
    analyzed_impact = sum(broken_urls[url]['count'] for url in analyzed_urls if url in broken_urls)
    total_impact = sum(info['count'] for info in broken_urls.values())
    impact_pct = (analyzed_impact / total_impact * 100) if total_impact > 0 else 0
    
    st.markdown(f"📊 These links affect **{analyzed_impact:,}** pages ({impact_pct:.0f}% of total impact)")
    
    st.markdown("")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Review Suggestions", type="primary", use_container_width=True):
            st.session_state.show_bulk_ai_modal = False
            st.session_state.bulk_ai_just_completed = False
            st.toast(f"✅ Analyzed {analyzed_count} URLs - review suggestions below", icon="✅")
            st.rerun()
    with col2:
        if remaining > 0:
            next_batch = min(remaining, 100)
            if st.button(f"Analyze Next {next_batch}", use_container_width=True):
                st.session_state.bulk_ai_just_completed = False
                st.rerun()
        else:
            st.button("All URLs Analyzed ✓", use_container_width=True, disabled=True)


def render_compact_url_row(url: str, info: Dict, decision: Dict, domain: str):
    """Render a compact inline-editable URL row with 4 columns"""
    import html
    
    # Determine current state
    has_ai = bool(decision['ai_action'])
    is_approved = bool(decision['approved_action'])
    is_editing = st.session_state.editing_url == url
    
    # Display URL - strip protocol for display only, keep full URL for link
    display_url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    
    # Truncate display only if very long (80+ chars)
    if len(display_url) > 80:
        truncated_url = display_url[:77] + "..."
    else:
        truncated_url = display_url
    
    # Escape for HTML safety
    url_escaped = html.escape(url)
    truncated_url_escaped = html.escape(truncated_url)
    
    # Build the row - 4 column layout: Broken URL | Fix | Approve | Edit
    with st.container():
        cols = st.columns([4, 4, 1, 0.8])
        
        with cols[0]:
            # Broken URL with link
            st.markdown(f"<a href='{url_escaped}' target='_blank' style='color: #0f172a; text-decoration: none; font-size: 0.85rem; word-break: break-all;' title='{url_escaped}'>{truncated_url_escaped}</a>", unsafe_allow_html=True)
        
        with cols[1]:
            # Fix/Solution column
            if is_approved:
                if decision['approved_action'] == 'remove':
                    st.markdown(f"<span style='color: #059669; font-size: 0.85rem;'>✅ Remove link</span>", unsafe_allow_html=True)
                elif decision['approved_action'] == 'replace':
                    fix_url = decision['approved_fix']
                    fix_url_escaped = html.escape(fix_url)
                    fix_display = fix_url.replace('https://', '').replace('http://', '').replace('www.', '')
                    if len(fix_display) > 50:
                        fix_display = fix_display[:47] + "..."
                    fix_display_escaped = html.escape(fix_display)
                    st.markdown(f"<span style='color: #059669; font-size: 0.85rem;'>✅ </span><a href='{fix_url_escaped}' target='_blank' style='color: #059669; text-decoration: none; font-size: 0.85rem;' title='{fix_url_escaped}'>{fix_display_escaped}</a>", unsafe_allow_html=True)
                elif decision['approved_action'] == 'ignore':
                    st.markdown(f"<span style='color: #64748b; font-size: 0.85rem;'>⏭️ Ignored</span>", unsafe_allow_html=True)
            elif has_ai:
                if decision['ai_action'] == 'remove':
                    st.markdown(f"<span style='color: #ca8a04; font-size: 0.85rem;'>🤖 Remove link</span>", unsafe_allow_html=True)
                else:
                    ai_url = decision['ai_suggestion']
                    ai_url_escaped = html.escape(ai_url)
                    ai_display = ai_url.replace('https://', '').replace('http://', '').replace('www.', '')
                    if len(ai_display) > 50:
                        ai_display = ai_display[:47] + "..."
                    ai_display_escaped = html.escape(ai_display)
                    st.markdown(f"<span style='color: #ca8a04; font-size: 0.85rem;'>🤖 </span><a href='{ai_url_escaped}' target='_blank' style='color: #ca8a04; text-decoration: none; font-size: 0.85rem;' title='{ai_url_escaped}'>{ai_display_escaped}</a>", unsafe_allow_html=True)
            else:
                st.markdown(f"<span style='color: #94a3b8; font-size: 0.85rem;'>—</span>", unsafe_allow_html=True)
        
        with cols[2]:
            # Approve column - checkbox style
            if is_approved:
                # Already approved - show green check
                st.markdown("<span style='color: #22c55e; font-size: 1.2rem;'>✓</span>", unsafe_allow_html=True)
            elif has_ai:
                # Has AI suggestion - show clickable approve checkbox
                if st.button("☐", key=f"approve_ai_{url}", help="Approve AI suggestion"):
                    decision['approved_action'] = decision['ai_action']
                    decision['approved_fix'] = decision['ai_suggestion']
                    st.toast("✅ Approved", icon="✅")
                    st.rerun()
            else:
                # No suggestion yet - show dash
                st.markdown("<span style='color: #cbd5e1;'>—</span>", unsafe_allow_html=True)
        
        with cols[3]:
            # Edit column
            if is_editing:
                if st.button("✕", key=f"close_{url}", help="Close"):
                    st.session_state.editing_url = None
                    st.rerun()
            elif is_approved:
                if st.button("📝", key=f"edit_{url}", help="Edit fix"):
                    st.session_state.editing_url = url
                    st.rerun()
            else:
                if st.button("📝", key=f"edit_{url}", help="Set fix"):
                    st.session_state.editing_url = url
                    st.rerun()
        
        # Expanded edit section (if this row is being edited)
        if is_editing:
            render_inline_edit(url, info, decision, domain)


def render_inline_edit(url: str, info: Dict, decision: Dict, domain: str):
    """Render compact inline edit form"""
    
    with st.container():
        st.markdown("---")
        
        # Show context info
        page_count = info['count']
        status_code = info['status_code']
        link_type = "Internal" if info['is_internal'] else "External"
        st.markdown(f"<span style='color: #64748b; font-size: 0.85rem;'>📊 Affects **{page_count}** page(s) · Status: **{status_code}** · {link_type}</span>", unsafe_allow_html=True)
        
        st.markdown("")  # Small spacer
        
        # Quick action row
        action_cols = st.columns([1, 1, 1, 1, 1])
        
        with action_cols[0]:
            if st.button("🗑️ Remove", key=f"quick_remove_{url}", use_container_width=True, help="Remove the link, keep anchor text"):
                decision['approved_action'] = 'remove'
                decision['approved_fix'] = ''
                st.session_state.editing_url = None
                st.toast("✅ Set to Remove", icon="✅")
                st.rerun()
        
        with action_cols[1]:
            if st.button("⏭️ Ignore", key=f"quick_ignore_{url}", use_container_width=True, help="Skip this URL"):
                decision['approved_action'] = 'ignore'
                decision['approved_fix'] = ''
                st.session_state.editing_url = None
                st.toast("✅ Ignored", icon="✅")
                st.rerun()
        
        with action_cols[2]:
            # Single AI button
            has_ai_key = bool(st.session_state.anthropic_key) or (AGENT_MODE_API_KEY and st.session_state.ai_suggestions_remaining > 0)
            if has_ai_key:
                if st.button("🤖 Get AI", key=f"quick_ai_{url}", use_container_width=True, help="Get AI suggestion"):
                    api_key = st.session_state.anthropic_key or AGENT_MODE_API_KEY
                    with st.spinner("AI analyzing..."):
                        result = get_ai_suggestion(url, info, domain, api_key)
                        decision['ai_action'] = result['action']
                        decision['ai_suggestion'] = result['url'] or ''
                        decision['ai_notes'] = result['notes']
                        if not st.session_state.anthropic_key:
                            st.session_state.ai_suggestions_remaining = max(0, st.session_state.ai_suggestions_remaining - 1)
                    st.rerun()
            else:
                st.button("🤖 Get AI", key=f"quick_ai_{url}", use_container_width=True, disabled=True, help="Add API key first")
        
        with action_cols[3]:
            pass  # Spacer
        
        with action_cols[4]:
            if st.button("❌ Close", key=f"close_edit_{url}", use_container_width=True):
                st.session_state.editing_url = None
                st.rerun()
        
        # Show AI suggestion if available
        if decision['ai_action']:
            st.markdown("**AI Suggestion:**")
            if decision['ai_action'] == 'remove':
                st.info(f"🗑️ **Remove link** — {decision['ai_notes'][:100]}..." if len(decision.get('ai_notes', '')) > 100 else f"🗑️ **Remove link** — {decision.get('ai_notes', 'No suitable replacement found')}")
            else:
                st.success(f"🔗 **Replace with:** `{decision['ai_suggestion']}`")
                if decision.get('ai_notes'):
                    st.caption(decision['ai_notes'][:150] + "..." if len(decision['ai_notes']) > 150 else decision['ai_notes'])
            
            if st.button("✅ Accept AI Suggestion", key=f"accept_ai_{url}", type="primary", use_container_width=True):
                decision['approved_action'] = decision['ai_action']
                decision['approved_fix'] = decision['ai_suggestion']
                st.session_state.editing_url = None
                st.toast("✅ Approved", icon="✅")
                st.rerun()
        
        # Manual replacement input
        st.markdown("**Or enter replacement URL:**")
        manual_cols = st.columns([4, 1])
        
        # Get base URL from domain - only for internal links
        is_internal = info.get('is_internal', True)
        if is_internal and domain:
            base_url = f"https://{domain}/"
            placeholder = f"{base_url}new-page-path"
        else:
            base_url = ""
            placeholder = "https://example.com/new-page"
        
        # Determine default value - use existing manual_fix, or base URL if empty
        existing_value = decision.get('manual_fix', '')
        default_value = existing_value if existing_value else base_url
        
        with manual_cols[0]:
            manual_url = st.text_input(
                "Replacement URL",
                value=default_value,
                placeholder=placeholder,
                key=f"manual_input_{url}",
                label_visibility="collapsed"
            )
        with manual_cols[1]:
            # Check if URL has content
            has_valid_url = manual_url and len(manual_url.strip()) > 10 and manual_url.startswith('http')
            if st.button("✓ Save", key=f"save_manual_{url}", type="primary", use_container_width=True):
                if has_valid_url:
                    decision['approved_action'] = 'replace'
                    decision['approved_fix'] = manual_url
                    decision['manual_fix'] = manual_url
                    st.session_state.editing_url = None
                    st.toast("✅ Saved", icon="✅")
                    st.rerun()
                else:
                    st.toast("⚠️ Enter a full URL path", icon="⚠️")
        
        # Reset button if already approved
        if decision['approved_action']:
            if st.button("↩️ Reset to Pending", key=f"reset_{url}"):
                decision['approved_action'] = ''
                decision['approved_fix'] = ''
                st.rerun()
        
        st.markdown("---")


def render_wordpress_section():
    """Render WordPress connection section"""
    st.markdown('<p class="section-header">WordPress Connection</p>', unsafe_allow_html=True)
    
    if st.session_state.wp_connected:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.success("✅ Connected to WordPress")
        with col2:
            if st.button("Disconnect"):
                if st.session_state.wp_client:
                    st.session_state.wp_client.close()
                st.session_state.wp_connected = False
                st.session_state.wp_client = None
                st.rerun()
    else:
        st.markdown("""
        Required to publish updates directly to your site. This connection will allow the tool to apply fixes directly to your WordPress site via the REST API. This step takes less than 3 minutes to setup in WordPress.
        """)
        
        # Setup instructions dropdown
        with st.expander("📋 **How to Get Your Application Password** (required)", expanded=False):
            st.markdown("""
            WordPress Application Passwords let external tools like Screaming Fixes access your site securely. 
            **You'll need a WordPress account with Administrator privileges.**
            
            ---
            
            **Step 1:** Log into your WordPress Admin dashboard
            
            **Step 2:** Go to **Users → Profile** (or click your name in the top-right corner)
            
            **Step 3:** Scroll down to the **Application Passwords** section
            
            **Step 4:** In the "New Application Password Name" field, enter: `Screaming Fixes`
            
            **Step 5:** Click **Add New Application Password**
            
            **Step 6:** Copy the generated password (it looks like `xxxx xxxx xxxx xxxx xxxx xxxx`)
            - ⚠️ **Important:** You'll only see this password once! Copy it now.
            - The spaces in the password are fine — paste it exactly as shown.
            
            ---
            
            **Troubleshooting:**
            - **Don't see Application Passwords?** You need WordPress 5.6 or higher, or install the [Application Passwords plugin](https://wordpress.org/plugins/application-passwords/).
            - **Getting 401 errors?** Make sure your username is correct and you're using an Application Password (not your regular login password).
            - **Getting 403 errors?** Your user account may not have permission to edit posts. Ask a site admin for help.
            """)
        
        col1, col2 = st.columns(2)
        with col1:
            site_url = st.text_input(
                "Site URL",
                placeholder="https://your-site.com",
                help="Your WordPress site URL (without /wp-json)"
            )
        with col2:
            username = st.text_input(
                "Username",
                placeholder="admin",
                help="Your WordPress admin username"
            )
        
        app_password = st.text_input(
            "Application Password ⚠️ This is NOT your WordPress login password — see instructions above",
            type="password",
            placeholder="xxxx xxxx xxxx xxxx xxxx xxxx",
            help="Generate this in WordPress under Users → Profile → Application Passwords. This is a separate password specifically for API access."
        )
        
        if st.button("🔌 Connect to WordPress", type="primary"):
            if not all([site_url, username, app_password]):
                st.error("Please fill in all fields")
            elif not WP_AVAILABLE:
                st.error("WordPress client not available. Check that wordpress_client.py is present.")
            else:
                with st.spinner("Connecting..."):
                    try:
                        client = WordPressClient(site_url, username, app_password)
                        result = client.test_connection()
                        
                        if result["success"]:
                            st.session_state.wp_connected = True
                            st.session_state.wp_client = client
                            st.success(f"✅ {result['message']}")
                            st.rerun()
                        else:
                            st.error(result["message"])
                    except Exception as e:
                        st.error(f"Connection failed: {str(e)}")
        
        st.markdown("""
        <div class="privacy-notice" style="margin-top: 1rem;">
            🔒 <strong>Your credentials are safe</strong> — stored in your browser session only. Nothing is saved to any database. All data is cleared when you close the tab.
        </div>
        """, unsafe_allow_html=True)


def create_gsheets_report() -> bytes:
    """Create a professional Excel report optimized for Google Sheets upload"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
    from io import BytesIO
    
    decisions = st.session_state.decisions
    broken_urls = st.session_state.broken_urls
    domain = st.session_state.domain or "Unknown"
    
    wb = Workbook()
    
    # =========================================================================
    # TAB 1: SUMMARY (Executive Overview)
    # =========================================================================
    summary = wb.active
    summary.title = "Summary"
    
    # Styling
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    metric_font = Font(bold=True, size=24, color="1a365d")
    label_font = Font(size=11, color="64748b")
    section_font = Font(bold=True, size=14, color="1a365d")
    
    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )
    
    # Title
    summary['A1'] = "🔗 Broken Links Audit Report"
    summary['A1'].font = Font(bold=True, size=20, color="1a365d")
    summary.merge_cells('A1:F1')
    
    # Metadata
    summary['A2'] = f"Site: {domain}"
    summary['A2'].font = Font(size=11, color="64748b")
    summary['D2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    summary['D2'].font = Font(size=11, color="64748b")
    summary.merge_cells('A2:C2')
    summary.merge_cells('D2:F2')
    
    # Calculate stats
    total_broken = len(broken_urls)
    total_fixed = sum(1 for d in decisions.values() if d['approved_action'] in ['remove', 'replace'])
    total_ignored = sum(1 for d in decisions.values() if d['approved_action'] == 'ignore')
    total_pending = sum(1 for d in decisions.values() if not d['approved_action'])
    total_remove = sum(1 for d in decisions.values() if d['approved_action'] == 'remove')
    total_replace = sum(1 for d in decisions.values() if d['approved_action'] == 'replace')
    
    total_pages_affected = sum(info['count'] for info in broken_urls.values())
    internal_count = sum(1 for info in broken_urls.values() if info['is_internal'])
    external_count = total_broken - internal_count
    
    # Key Metrics Section
    summary['A4'] = "KEY METRICS"
    summary['A4'].font = section_font
    summary.merge_cells('A4:F4')
    
    # Metrics row
    metrics = [
        ("Total Broken", total_broken),
        ("Fixed", total_fixed),
        ("Ignored", total_ignored),
        ("Pending", total_pending),
        ("Pages Affected", total_pages_affected),
    ]
    
    for i, (label, value) in enumerate(metrics):
        col = get_column_letter(i + 1)
        summary[f'{col}5'] = value
        summary[f'{col}5'].font = metric_font
        summary[f'{col}5'].alignment = Alignment(horizontal='center')
        summary[f'{col}6'] = label
        summary[f'{col}6'].font = label_font
        summary[f'{col}6'].alignment = Alignment(horizontal='center')
    
    # Breakdown Section
    summary['A8'] = "BREAKDOWN"
    summary['A8'].font = section_font
    
    breakdown_data = [
        ["Category", "Count", "Percentage"],
        ["Internal Links", internal_count, f"{internal_count/total_broken*100:.1f}%" if total_broken > 0 else "0%"],
        ["External Links", external_count, f"{external_count/total_broken*100:.1f}%" if total_broken > 0 else "0%"],
        ["", "", ""],
        ["Links Removed", total_remove, f"{total_remove/total_broken*100:.1f}%" if total_broken > 0 else "0%"],
        ["Links Replaced", total_replace, f"{total_replace/total_broken*100:.1f}%" if total_broken > 0 else "0%"],
        ["Ignored", total_ignored, f"{total_ignored/total_broken*100:.1f}%" if total_broken > 0 else "0%"],
        ["Pending Review", total_pending, f"{total_pending/total_broken*100:.1f}%" if total_broken > 0 else "0%"],
    ]
    
    for row_idx, row_data in enumerate(breakdown_data, start=9):
        for col_idx, value in enumerate(row_data, start=1):
            cell = summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 9:  # Header row
                cell.fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
                cell.font = Font(bold=True)
    
    # Status codes breakdown
    status_counts = {}
    for info in broken_urls.values():
        code = info['status_code']
        status_counts[code] = status_counts.get(code, 0) + 1
    
    summary['A19'] = "STATUS CODES"
    summary['A19'].font = section_font
    
    status_row = 20
    summary.cell(row=status_row, column=1, value="Status Code").font = Font(bold=True)
    summary.cell(row=status_row, column=2, value="Count").font = Font(bold=True)
    summary.cell(row=status_row, column=3, value="Meaning").font = Font(bold=True)
    
    status_meanings = {
        400: "Bad Request",
        401: "Unauthorized", 
        403: "Forbidden",
        404: "Not Found",
        410: "Gone (Intentionally Removed)",
        500: "Server Error",
        502: "Bad Gateway",
        503: "Service Unavailable"
    }
    
    for i, (code, count) in enumerate(sorted(status_counts.items()), start=1):
        row = status_row + i
        summary.cell(row=row, column=1, value=code)
        summary.cell(row=row, column=2, value=count)
        summary.cell(row=row, column=3, value=status_meanings.get(code, "Other Error"))
    
    # Column widths
    summary.column_dimensions['A'].width = 18
    summary.column_dimensions['B'].width = 12
    summary.column_dimensions['C'].width = 12
    summary.column_dimensions['D'].width = 25
    summary.column_dimensions['E'].width = 15
    summary.column_dimensions['F'].width = 15
    
    # =========================================================================
    # TAB 2: FIXES (Action Log)
    # =========================================================================
    fixes = wb.create_sheet("Fixes")
    
    # Header
    fix_headers = ["Broken URL", "Status Code", "Type", "Action", "New URL", "Pages Affected", "Anchor Text", "AI Notes"]
    for col, header in enumerate(fix_headers, start=1):
        cell = fixes.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 2
    for url, decision in decisions.items():
        if not decision['approved_action']:
            continue  # Skip pending
            
        info = broken_urls.get(url, {})
        
        # Determine action text
        if decision['approved_action'] == 'remove':
            action_text = "🗑️ Remove Link"
            new_url = ""
        elif decision['approved_action'] == 'replace':
            action_text = "🔗 Replace"
            new_url = decision['approved_fix']
        elif decision['approved_action'] == 'ignore':
            action_text = "⏭️ Ignored"
            new_url = ""
        else:
            action_text = decision['approved_action']
            new_url = decision['approved_fix']
        
        anchors = ", ".join(info.get('anchors', [])[:3])
        if len(info.get('anchors', [])) > 3:
            anchors += f" (+{len(info['anchors'])-3} more)"
        
        fixes.cell(row=row, column=1, value=url)
        fixes.cell(row=row, column=2, value=info.get('status_code', ''))
        fixes.cell(row=row, column=3, value="Internal" if info.get('is_internal') else "External")
        fixes.cell(row=row, column=4, value=action_text)
        fixes.cell(row=row, column=5, value=new_url)
        fixes.cell(row=row, column=6, value=info.get('count', 0))
        fixes.cell(row=row, column=7, value=anchors)
        fixes.cell(row=row, column=8, value=decision.get('ai_notes', ''))
        
        # Color code rows
        if decision['approved_action'] == 'remove':
            fill = PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid")
        elif decision['approved_action'] == 'replace':
            fill = PatternFill(start_color="f0fdf4", end_color="f0fdf4", fill_type="solid")
        else:
            fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
        
        for col in range(1, 9):
            fixes.cell(row=row, column=col).fill = fill
            fixes.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # Column widths
    fixes.column_dimensions['A'].width = 50
    fixes.column_dimensions['B'].width = 12
    fixes.column_dimensions['C'].width = 10
    fixes.column_dimensions['D'].width = 15
    fixes.column_dimensions['E'].width = 50
    fixes.column_dimensions['F'].width = 12
    fixes.column_dimensions['G'].width = 30
    fixes.column_dimensions['H'].width = 40
    
    # Freeze header row
    fixes.freeze_panes = 'A2'
    
    # =========================================================================
    # TAB 3: DETAILED LOG (Every source page)
    # =========================================================================
    details = wb.create_sheet("Detailed Log")
    
    # Header
    detail_headers = ["Source Page", "Post ID", "Broken URL", "Anchor Text", "Action", "New URL", "Link Position"]
    for col, header in enumerate(detail_headers, start=1):
        cell = details.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data - one row per source page affected
    row = 2
    for url, decision in decisions.items():
        if not decision['approved_action'] or decision['approved_action'] == 'ignore':
            continue
            
        info = broken_urls.get(url, {})
        
        # Get action and new URL
        if decision['approved_action'] == 'remove':
            action_text = "Remove"
            new_url = ""
        else:
            action_text = "Replace"
            new_url = decision['approved_fix']
        
        # One row per source page
        for source in info.get('sources', []):
            post_id = st.session_state.post_id_cache.get(source, "")
            anchor = info.get('anchors', [''])[0] if info.get('anchors') else ''
            
            details.cell(row=row, column=1, value=source)
            details.cell(row=row, column=2, value=post_id)
            details.cell(row=row, column=3, value=url)
            details.cell(row=row, column=4, value=anchor)
            details.cell(row=row, column=5, value=action_text)
            details.cell(row=row, column=6, value=new_url)
            details.cell(row=row, column=7, value=info.get('link_positions', ['Content'])[0] if info.get('link_positions') else 'Content')
            
            for col in range(1, 8):
                details.cell(row=row, column=col).border = thin_border
            
            row += 1
    
    # Column widths
    details.column_dimensions['A'].width = 50
    details.column_dimensions['B'].width = 10
    details.column_dimensions['C'].width = 50
    details.column_dimensions['D'].width = 25
    details.column_dimensions['E'].width = 12
    details.column_dimensions['F'].width = 50
    details.column_dimensions['G'].width = 15
    
    # Freeze header row
    details.freeze_panes = 'A2'
    
    # =========================================================================
    # TAB 4: PENDING (Still needs review)
    # =========================================================================
    pending = wb.create_sheet("Pending Review")
    
    # Header
    pending_headers = ["Broken URL", "Status Code", "Type", "Pages Affected", "Anchor Text", "AI Suggestion"]
    for col, header in enumerate(pending_headers, start=1):
        cell = pending.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Data
    row = 2
    for url, decision in decisions.items():
        if decision['approved_action']:
            continue  # Skip approved
            
        info = broken_urls.get(url, {})
        anchors = ", ".join(info.get('anchors', [])[:2])
        
        ai_suggestion = ""
        if decision.get('ai_action'):
            if decision['ai_action'] == 'remove':
                ai_suggestion = "AI suggests: Remove"
            else:
                ai_suggestion = f"AI suggests: {decision.get('ai_suggestion', '')}"
        
        pending.cell(row=row, column=1, value=url)
        pending.cell(row=row, column=2, value=info.get('status_code', ''))
        pending.cell(row=row, column=3, value="Internal" if info.get('is_internal') else "External")
        pending.cell(row=row, column=4, value=info.get('count', 0))
        pending.cell(row=row, column=5, value=anchors)
        pending.cell(row=row, column=6, value=ai_suggestion)
        
        for col in range(1, 7):
            pending.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # Column widths
    pending.column_dimensions['A'].width = 50
    pending.column_dimensions['B'].width = 12
    pending.column_dimensions['C'].width = 10
    pending.column_dimensions['D'].width = 12
    pending.column_dimensions['E'].width = 30
    pending.column_dimensions['F'].width = 40
    
    pending.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render_gsheets_export_modal():
    """Render Google Sheets export modal"""
    st.markdown("---")
    st.markdown("### 📊 Export for Google Sheets")
    
    st.markdown("""
    Generate a **professional SEO audit report** in Excel format, optimized for Google Sheets upload.
    
    **Your report includes:**
    - 📈 **Summary** — Executive overview with key metrics and charts
    - ✅ **Fixes** — All approved actions (remove/replace) with details  
    - 📋 **Detailed Log** — Every affected page with Post IDs
    - ⏳ **Pending Review** — URLs still needing attention
    """)
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📥 Generate Report", type="primary", use_container_width=True):
            with st.spinner("Creating your report..."):
                try:
                    xlsx_data = create_gsheets_report()
                    
                    # Track export
                    track_event("export", {
                        "format": "xlsx_gsheets",
                        "total_urls": len(st.session_state.broken_urls)
                    })
                    
                    st.session_state.gsheets_export_data = xlsx_data
                    st.success("✅ Report ready!")
                except Exception as e:
                    st.error(f"Error creating report: {str(e)}")
    
    with col2:
        if st.button("❌ Cancel", use_container_width=True):
            st.session_state.show_gsheets_export = False
            st.rerun()
    
    # Show download button if report is ready
    if st.session_state.get('gsheets_export_data'):
        domain = st.session_state.domain or "site"
        filename = f"broken_links_audit_{domain}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        st.download_button(
            "⬇️ Download Excel Report",
            data=st.session_state.gsheets_export_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.markdown("""
        **Next steps:**
        1. Download the report above
        2. Go to [Google Drive](https://drive.google.com)
        3. Click **New → File upload** and select the Excel file
        4. Right-click the file → **Open with → Google Sheets**
        
        💡 **Tip:** Share the Google Sheet link with your team or clients for collaboration!
        """)


def render_export_section():
    """Render export section"""
    decisions = st.session_state.decisions
    broken_urls = st.session_state.broken_urls
    
    approved = [(url, d) for url, d in decisions.items() if d['approved_action'] and d['approved_action'] != 'ignore']
    has_approved = len(approved) > 0
    
    st.markdown('<p class="section-header">Export & Apply</p>', unsafe_allow_html=True)
    
    # Status message based on approved count
    if has_approved:
        st.markdown(f"✅ **{len(approved)} fixes approved** and ready")
    else:
        st.markdown("No fixes approved yet. Approve fixes above to enable export.")
    
    # Export buttons - 3 columns now
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📊 Export for Google Sheets", use_container_width=True, disabled=not has_approved, type="primary"):
            st.session_state.show_gsheets_export = True
    
    with col2:
        if st.button("📥 Export CSV", use_container_width=True, disabled=not has_approved):
            export_data = create_export_data()
            
            # Track export (counts only, no URLs)
            remove_count = sum(1 for d in export_data if d['action'] == 'remove')
            replace_count = sum(1 for d in export_data if d['action'] == 'replace')
            track_event("export", {
                "format": "csv",
                "total_fixes": len(export_data),
                "remove_count": remove_count,
                "replace_count": replace_count
            })
            
            csv_data = pd.DataFrame(export_data).to_csv(index=False)
            st.download_button(
                "Download CSV",
                data=csv_data,
                file_name=f"screaming_fixes_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    with col3:
        if st.button("📥 Export JSON", use_container_width=True, disabled=not has_approved):
            export_data = create_export_data()
            
            # Track export (counts only, no URLs)
            remove_count = sum(1 for d in export_data if d['action'] == 'remove')
            replace_count = sum(1 for d in export_data if d['action'] == 'replace')
            track_event("export", {
                "format": "json",
                "total_fixes": len(export_data),
                "remove_count": remove_count,
                "replace_count": replace_count
            })
            
            json_data = json.dumps(export_data, indent=2)
            st.download_button(
                "Download JSON",
                data=json_data,
                file_name=f"screaming_fixes_{datetime.now().strftime('%Y%m%d')}.json",
                mime="application/json",
                use_container_width=True
            )
    
    # Google Sheets Export Modal
    if st.session_state.get('show_gsheets_export'):
        render_gsheets_export_modal()
    
    # WordPress apply section - only show after connected
    if not st.session_state.wp_connected:
        # Don't show this section until WordPress is connected
        return
    
    st.markdown("---")
    st.markdown("**Apply directly to WordPress:**")
    
    if not has_approved:
        st.info("Approve some fixes above first")
        return
    
    # Count how many source pages need fixes
    source_pages_to_fix = set()
    for url, decision in approved:
        for source in broken_urls[url]['sources']:
            source_pages_to_fix.add(source)
    
    pages_to_fix = len(source_pages_to_fix)
    has_post_ids = st.session_state.has_post_ids
    selected_mode = st.session_state.get('selected_mode', 'quick_start')
    
    # Check if we have post_ids for all pages
    if has_post_ids:
        pages_with_ids = sum(1 for s in source_pages_to_fix if s in st.session_state.post_id_cache)
        all_have_ids = pages_with_ids == pages_to_fix
    else:
        all_have_ids = False
        pages_with_ids = 0
    
    # Handle based on selected mode
    if selected_mode == 'full':
        # Full Mode selected
        if all_have_ids:
            st.success(f"🚀 **Full Mode:** All {pages_to_fix} pages have Post IDs — ready for maximum speed!")
            render_wordpress_execute_ui(approved, source_pages_to_fix)
        else:
            st.error("🚀 **Full Mode:** Post IDs required but not found in your CSV.")
            st.markdown("**To use Full Mode:**")
            st.markdown("1. Set up Screaming Frog to extract Post IDs (see instructions above)")
            st.markdown("2. Re-crawl your site")
            st.markdown("3. Re-upload your CSV")
            st.markdown("")
            st.info("💡 **Or switch to Quick Start Mode** to fix up to 25 pages right now!")
    
    else:
        # Quick Start Mode selected
        if pages_to_fix <= AGENT_MODE_LIMIT:
            # Within limit - can fix all
            st.info(f"🤖 **Quick Start Mode:** {pages_to_fix} pages to update")
            
            # Sample check if not done
            if not st.session_state.post_id_check_done:
                st.markdown("Before applying, let's verify your site supports automatic Post ID discovery.")
                if st.button("🔍 Run Compatibility Check", use_container_width=True):
                    run_post_id_sample_check(list(source_pages_to_fix)[:3])
            else:
                if st.session_state.post_id_check_passed:
                    st.success("✅ Compatibility check passed! Your site supports automatic Post ID lookup.")
                    render_wordpress_execute_ui(approved, source_pages_to_fix)
                else:
                    st.error("❌ Could not find Post IDs automatically on your site.")
                    st.info("💡 Try setting up Full Mode for more reliable Post ID handling.")
        
        else:
            # Over limit - offer test run
            pages_over = pages_to_fix - AGENT_MODE_LIMIT
            
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #e0f2fe 0%, #bae6fd 100%); padding: 1.25rem; border-radius: 10px; margin-bottom: 1rem; border: 1px solid #7dd3fc;">
                <div style="font-weight: 600; color: #0c4a6e; margin-bottom: 0.5rem;">🤖 Quick Start Mode: Test Run</div>
                <div style="color: #0369a1; font-size: 0.95rem;">
                    You have {pages_to_fix} pages, but Quick Start Mode handles {AGENT_MODE_LIMIT} at a time.<br>
                    <strong>Let's fix the first {AGENT_MODE_LIMIT} pages</strong> so you can see the tool in action!
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Select first 25 pages
            test_run_pages = list(source_pages_to_fix)[:AGENT_MODE_LIMIT]
            
            # Filter approved fixes to only those affecting test run pages
            test_run_approved = []
            for url, decision in approved:
                sources_in_test = [s for s in broken_urls[url]['sources'] if s in test_run_pages]
                if sources_in_test:
                    test_run_approved.append((url, decision))
            
            st.markdown(f"**Test run:** {len(test_run_approved)} broken URLs across {len(test_run_pages)} pages")
            
            # Sample check if not done
            if not st.session_state.post_id_check_done:
                if st.button("🔍 Run Compatibility Check", use_container_width=True, key="test_run_check"):
                    run_post_id_sample_check(test_run_pages[:3])
            else:
                if st.session_state.post_id_check_passed:
                    st.success("✅ Compatibility check passed!")
                    render_wordpress_execute_ui(test_run_approved, set(test_run_pages), is_test_run=True)
                else:
                    st.error("❌ Automatic Post ID lookup failed.")
                    st.info("💡 Try setting up Full Mode for more reliable Post ID handling.")


def run_post_id_sample_check(sample_urls: List[str]):
    """Run sample check on a few URLs to verify Post ID discovery works"""
    client = st.session_state.wp_client
    
    with st.status("Checking compatibility...", expanded=True) as status:
        found = 0
        failed = 0
        
        for url in sample_urls:
            st.write(f"Testing: `{url[:60]}...`")
            post_id = client.find_post_id_by_url(url)
            
            if post_id:
                st.write(f"  ✅ Found Post ID: {post_id}")
                found += 1
                st.session_state.post_id_cache[url] = post_id
            else:
                st.write(f"  ❌ Post ID not found")
                failed += 1
        
        st.session_state.post_id_check_done = True
        
        # Pass if at least half succeeded
        if found >= len(sample_urls) / 2:
            st.session_state.post_id_check_passed = True
            status.update(label="✅ Compatibility check passed!", state="complete")
        else:
            st.session_state.post_id_check_passed = False
            status.update(label="❌ Compatibility check failed", state="error")
    
    st.rerun()


def render_large_dataset_instructions():
    """Show instructions for handling large datasets"""
    st.markdown("---")
    st.markdown("### 📋 How to Add Post IDs to Your CSV")
    
    with st.expander("**Option 1: Screaming Frog Custom Extraction (Recommended)**", expanded=True):
        st.markdown("""
        Add Post IDs directly during your crawl using Screaming Frog's custom extraction:
        
        **Setup (one-time):**
        1. In Screaming Frog, go to **Configuration → Custom → Extraction**
        2. Click **Add**
        3. Set the name to `post_id`
        4. Change the dropdown from "XPath" to **Regex**
        5. Enter this pattern:
        ```
        <link[^>]+rel=["']shortlink["'][^>]+href=["'][^"']*\\?p=(\\d+)
        ```
        6. Click **OK**
        
        **Run your crawl:**
        1. Start a new crawl of your site
        2. After completion, go to **Bulk Export → Custom Extraction → post_id**
        3. Save this CSV file and upload it here
        
        *Most WordPress sites include a shortlink meta tag with the Post ID.*
        """)
    
    with st.expander("**Option 2: Manual Post ID Lookup**"):
        st.markdown("""
        Find Post IDs manually in WordPress Admin:
        
        1. Go to **WordPress Admin → Posts** (or Pages)
        2. Hover over the post/page title
        3. Look at the URL in your browser status bar: `post.php?post=12345`
        4. The number after `post=` is your Post ID
        
        *Or:* Edit the post/page and check the URL — it contains `post=12345`
        
        Add a `post_id` column to your CSV and re-upload.
        """)
    
    with st.expander("**Option 3: Export & Update Manually**"):
        st.markdown("""
        Export your approved fixes as CSV and apply them manually:
        
        1. Click **Export CSV** above
        2. Open in Excel/Google Sheets
        3. For each row, find the Post ID and apply the fix in WordPress
        
        *This bypasses the API entirely but requires manual work.*
        """)


def render_wordpress_execute_ui(approved: List, source_pages_to_fix: set, is_test_run: bool = False):
    """Render the execute UI when ready to apply fixes"""
    pages_to_fix = len(source_pages_to_fix)
    
    st.markdown("---")
    
    if is_test_run:
        st.markdown(f"""
        <div style="background: #eff6ff; padding: 0.75rem 1rem; border-radius: 8px; border: 1px solid #bfdbfe; margin-bottom: 1rem;">
            <span style="font-weight: 600; color: #1e40af;">🧪 Test Run:</span>
            <span style="color: #3730a3;">Fixing {len(approved)} broken URLs across {pages_to_fix} pages (first {AGENT_MODE_LIMIT} pages only)</span>
        </div>
        """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        button_label = "🧪 Run Test (25 pages)" if is_test_run else "🚀 Apply Fixes to WordPress"
        if st.button(button_label, type="primary", use_container_width=True):
            run_agent_fixes(approved, source_pages_to_fix, is_test_run)
    
    with col2:
        if st.session_state.get('wp_execute_results'):
            if st.button("📋 View Last Results", use_container_width=True):
                pass  # Results shown below
    
    # Show results if any
    render_wp_results(is_test_run)


def run_agent_fixes(approved: List, source_pages_to_fix: set, is_test_run: bool = False):
    """Run the agent to apply fixes with live view"""
    client = st.session_state.wp_client
    broken_urls = st.session_state.broken_urls
    
    # Build list of all fixes to apply (only for pages in source_pages_to_fix)
    fixes_to_apply = []
    for url, decision in approved:
        info = broken_urls[url]
        for source in info['sources']:
            # Only include if this source is in our target pages
            if source in source_pages_to_fix:
                fixes_to_apply.append({
                    'source_url': source,
                    'broken_url': url,
                    'action': decision['approved_action'],
                    'replacement_url': decision['approved_fix'] if decision['approved_action'] == 'replace' else '',
                    'post_id': st.session_state.post_id_cache.get(source)  # May be None
                })
    
    total_fixes = len(fixes_to_apply)
    results = []
    skipped_for_retry = []
    
    st.markdown("### 🤖 Applying Fixes")
    
    progress_bar = st.progress(0)
    status_container = st.container()
    
    with status_container:
        for i, fix in enumerate(fixes_to_apply):
            progress_bar.progress((i + 1) / total_fixes)
            
            with st.status(f"Fix {i+1} of {total_fixes}", expanded=True) as status:
                # Step 1: Get Post ID
                post_id = fix['post_id']
                
                if not post_id:
                    st.write(f"🔍 Finding Post ID for: `{fix['source_url'][:50]}...`")
                    post_id = client.find_post_id_by_url(fix['source_url'])
                    
                    if post_id:
                        st.write(f"   ✅ Found Post ID: {post_id}")
                        st.session_state.post_id_cache[fix['source_url']] = post_id
                    else:
                        # Check if it's a category/archive/tag page
                        source_lower = fix['source_url'].lower()
                        is_archive = any(x in source_lower for x in ['/category/', '/tag/', '/author/', '/page/', '/archive/'])
                        
                        if is_archive:
                            st.write(f"   ℹ️ This is an archive/category page (dynamically generated)")
                            st.write(f"   💡 Fix the links on individual posts — archive pages will update automatically")
                            msg = 'Archive/category page - fix individual posts instead'
                        else:
                            st.write(f"   ❌ Post ID not found — skipping")
                            msg = 'Post ID not found'
                        
                        results.append({
                            'source_url': fix['source_url'],
                            'broken_url': fix['broken_url'],
                            'action': fix['action'],
                            'status': 'skipped',
                            'message': msg
                        })
                        skipped_for_retry.append(fix)
                        status.update(label=f"⏭️ Skipped", state="error")
                        continue
                else:
                    st.write(f"📄 Post ID: {post_id} (from CSV)")
                
                # Step 2: Apply fix
                st.write(f"🔧 Applying: **{fix['action'].upper()}**")
                st.write(f"   Target: `{fix['broken_url'][:50]}...`")
                
                try:
                    if fix['action'] == 'remove':
                        result = client.remove_link(post_id, fix['broken_url'], dry_run=False)
                    else:
                        result = client.replace_link(post_id, fix['broken_url'], fix['replacement_url'], dry_run=False)
                    
                    if result['success']:
                        st.write(f"   ✅ {result['message']}")
                        results.append({
                            'source_url': fix['source_url'],
                            'broken_url': fix['broken_url'],
                            'action': fix['action'],
                            'status': 'success',
                            'message': result['message']
                        })
                        status.update(label=f"✅ Success", state="complete")
                    else:
                        # Provide helpful context for "URL not found" errors
                        if 'not found' in result['message'].lower():
                            st.write(f"   ❌ {result['message']}")
                            st.write(f"   💡 Common reasons: link may be in a widget, shortcode, custom field, or theme template")
                        else:
                            st.write(f"   ❌ {result['message']}")
                        
                        results.append({
                            'source_url': fix['source_url'],
                            'broken_url': fix['broken_url'],
                            'action': fix['action'],
                            'status': 'failed',
                            'message': result['message']
                        })
                        status.update(label=f"❌ Failed", state="error")
                        
                except Exception as e:
                    st.write(f"   ❌ Error: {str(e)}")
                    results.append({
                        'source_url': fix['source_url'],
                        'broken_url': fix['broken_url'],
                        'action': fix['action'],
                        'status': 'failed',
                        'message': str(e)
                    })
                    status.update(label=f"❌ Error", state="error")
            
            # Small delay to be nice to the server
            time.sleep(0.3)
    
    # Store results
    st.session_state.wp_execute_results = results
    st.session_state.wp_was_test_run = is_test_run
    
    # Track execution
    success = sum(1 for r in results if r['status'] == 'success')
    skipped = sum(1 for r in results if r['status'] == 'skipped')
    failed = sum(1 for r in results if r['status'] == 'failed')
    
    track_event("wordpress_apply", {
        "mode": "agent_test_run" if is_test_run else "agent",
        "total_fixes": len(results),
        "success": success,
        "skipped": skipped,
        "failed": failed
    })
    
    # Summary
    st.markdown("---")
    if is_test_run:
        st.markdown("### 🧪 Test Run Complete!")
    else:
        st.markdown("### 📊 Execution Complete")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("✅ Success", success)
    with col2:
        st.metric("⏭️ Skipped", skipped)
    with col3:
        st.metric("❌ Failed", failed)
    
    # Test run follow-up
    if is_test_run and success > 0:
        total_pages = st.session_state.source_pages_count
        remaining_pages = total_pages - AGENT_MODE_LIMIT
        
        st.markdown("---")
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%); padding: 1.25rem; border-radius: 10px; border: 1px solid #6ee7b7;">
            <div style="font-weight: 600; color: #065f46; font-size: 1.1rem; margin-bottom: 0.5rem;">✅ Test Run Successful!</div>
            <div style="color: #047857; margin-bottom: 1rem;">
                You've verified that Screaming Fixes works on your site. 
                <strong>{remaining_pages} more pages</strong> are waiting to be fixed.
            </div>
            <div style="font-weight: 600; color: #065f46; margin-bottom: 0.5rem;">Next Steps:</div>
            <div style="color: #047857;">
                1. Set up <strong>Fast Mode</strong> in Screaming Frog (2 min setup)<br>
                2. Re-crawl your site to capture Post IDs<br>
                3. Re-upload your CSV — all {total_pages} pages will be ready to fix
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        with st.expander("📋 Set up Fast Mode now", expanded=True):
            render_post_id_extraction_guide()
    
    # Handle skipped pages
    if skipped_for_retry:
        st.markdown("---")
        st.warning(f"**{len(skipped_for_retry)} pages** were skipped because Post ID could not be found.")
        
        with st.expander("Enter Post IDs manually for skipped pages"):
            for fix in skipped_for_retry:
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.text(fix['source_url'][:60] + "..." if len(fix['source_url']) > 60 else fix['source_url'])
                with col2:
                    post_id_input = st.text_input(
                        "Post ID",
                        key=f"manual_postid_{fix['source_url']}",
                        label_visibility='collapsed',
                        placeholder="Post ID"
                    )
                    if post_id_input and post_id_input.isdigit():
                        st.session_state.post_id_cache[fix['source_url']] = int(post_id_input)
            
            if st.button("🔄 Retry Skipped Pages"):
                st.rerun()


def apply_fixes_to_wordpress(dry_run: bool = True):
    """Apply approved fixes to WordPress"""
    if not st.session_state.wp_client:
        st.error("WordPress not connected")
        return
    
    client = st.session_state.wp_client
    export_data = create_export_data()
    
    if not export_data:
        st.warning("No fixes to apply")
        return
    
    mode = "Preview" if dry_run else "Execution"
    
    with st.status(f"Running {mode}...", expanded=True) as status:
        results = []
        
        for i, fix in enumerate(export_data):
            st.write(f"Processing {i+1}/{len(export_data)}: `{fix['source_url'][:50]}...`")
            
            try:
                # Find post ID from source URL
                post_id = client.find_post_id_by_url(fix['source_url'])
                if not post_id:
                    results.append({
                        'source_url': fix['source_url'],
                        'broken_url': fix['broken_url'],
                        'action': fix['action'],
                        'status': 'skipped',
                        'message': 'Post ID not found'
                    })
                    continue
                
                # Apply fix (or preview)
                if fix['action'] == 'remove':
                    result = client.remove_link(post_id, fix['broken_url'], dry_run=dry_run)
                else:
                    result = client.replace_link(post_id, fix['broken_url'], fix['replacement_url'], dry_run=dry_run)
                
                results.append({
                    'source_url': fix['source_url'],
                    'broken_url': fix['broken_url'],
                    'action': fix['action'],
                    'status': 'success' if result['success'] else 'failed',
                    'message': result['message']
                })
            except Exception as e:
                results.append({
                    'source_url': fix['source_url'],
                    'broken_url': fix['broken_url'],
                    'action': fix['action'],
                    'status': 'failed',
                    'message': str(e)
                })
        
        # Store results
        if dry_run:
            st.session_state.wp_preview_results = results
            st.session_state.wp_preview_done = True
        else:
            st.session_state.wp_execute_results = results
            st.session_state.wp_preview_done = False  # Reset for next batch
        
        # Summary
        success = sum(1 for r in results if r['status'] == 'success')
        skipped = sum(1 for r in results if r['status'] == 'skipped')
        failed = sum(1 for r in results if r['status'] == 'failed')
        
        # Track WordPress apply (counts only, no URLs)
        track_event("wordpress_apply", {
            "mode": "preview" if dry_run else "execute",
            "total_fixes": len(results),
            "success": success,
            "skipped": skipped,
            "failed": failed
        })
        
        status.update(
            label=f"✅ {mode} complete: {success} successful, {skipped} skipped, {failed} failed",
            state="complete"
        )


def render_wp_results(is_test_run: bool = False):
    """Render WordPress results table"""
    # Show execute results if available, else preview results
    results = st.session_state.get('wp_execute_results') or st.session_state.get('wp_preview_results')
    
    if not results:
        return
    
    is_preview = st.session_state.get('wp_execute_results') is None
    was_test_run = st.session_state.get('wp_was_test_run', False)
    
    if was_test_run:
        st.markdown("### 🧪 Test Run Results")
    elif is_preview:
        st.markdown("### Preview Results")
    else:
        st.markdown("### Execution Results")
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total", len(results))
    with col2:
        st.metric("✅ Success", sum(1 for r in results if r['status'] == 'success'))
    with col3:
        st.metric("⏭️ Skipped", sum(1 for r in results if r['status'] == 'skipped'))
    with col4:
        st.metric("❌ Failed", sum(1 for r in results if r['status'] == 'failed'))
    
    # Results table (shortened for readability)
    df = pd.DataFrame([
        {
            'Source URL': r['source_url'][:50] + '...' if len(r['source_url']) > 50 else r['source_url'],
            'Broken URL': r['broken_url'][:40] + '...' if len(r['broken_url']) > 40 else r['broken_url'],
            'Action': r['action'].upper(),
            'Status': r['status'].upper(),
            'Message': r['message']
        }
        for r in results
    ])
    
    st.dataframe(df, use_container_width=True)
    
    # Full URLs expander for copy/paste
    with st.expander("📋 View Full URLs (for copy/paste)"):
        for i, r in enumerate(results):
            status_icon = "✅" if r['status'] == 'success' else ("⏭️" if r['status'] == 'skipped' else "❌")
            st.markdown(f"**{status_icon} Result {i+1}**")
            st.code(f"Source URL: {r['source_url']}\nBroken URL: {r['broken_url']}", language=None)
            st.markdown(f"Action: `{r['action'].upper()}` | Status: `{r['status'].upper()}` | {r['message']}")
            if i < len(results) - 1:
                st.markdown("---")
    
    # Download results
    csv_output = pd.DataFrame(results).to_csv(index=False)
    mode = "preview" if is_preview else "executed"
    
    st.download_button(
        label="📥 Download Results CSV",
        data=csv_output,
        file_name=f"screaming_fixes_{mode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mime="text/csv",
        use_container_width=True
    )


def create_export_data() -> List[Dict]:
    """Create export data from approved decisions"""
    decisions = st.session_state.decisions
    broken_urls = st.session_state.broken_urls
    
    export = []
    
    for url, decision in decisions.items():
        if not decision['approved_action'] or decision['approved_action'] == 'ignore':
            continue
        
        info = broken_urls[url]
        
        for source in info['sources']:
            export.append({
                'source_url': source,
                'broken_url': url,
                'status_code': info['status_code'],
                'is_internal': info['is_internal'],
                'action': decision['approved_action'],
                'replacement_url': decision['approved_fix'] if decision['approved_action'] == 'replace' else '',
            })
    
    return export


def create_rc_export_data() -> List[Dict]:
    """Create export data from approved redirect chain decisions"""
    decisions = st.session_state.rc_decisions
    redirects = st.session_state.rc_redirects
    
    export = []
    
    for key, decision in decisions.items():
        if not decision['approved_action']:
            continue
        
        info = redirects[key]
        
        for source in info['sources']:
            export.append({
                'source_url': source,
                'old_url': info['address'],
                'new_url': info['final_address'],
                'redirect_type': '302 (Temporary)' if info['is_temp_redirect'] else '301 (Permanent)',
                'hops': info['num_hops'],
                'action': 'replace',
            })
    
    return export


def render_task_switcher():
    """Render task type switcher dropdown"""
    # Count items in each task type
    broken_count = len(st.session_state.broken_urls) if st.session_state.broken_urls else 0
    broken_pending = sum(1 for d in st.session_state.decisions.values() if not d.get('approved_action')) if st.session_state.decisions else 0

    rc_count = len(st.session_state.rc_redirects) if st.session_state.rc_redirects else 0
    rc_pending = sum(1 for d in st.session_state.rc_decisions.values() if not d.get('approved_action')) if st.session_state.rc_decisions else 0

    iat_count = len(st.session_state.iat_images) if st.session_state.iat_images else 0
    iat_pending = sum(1 for d in st.session_state.iat_decisions.values() if not d.get('approved_action')) if st.session_state.iat_decisions else 0

    # Backlink Reclaim counts
    br_count = len(st.session_state.br_grouped_pages) if st.session_state.br_grouped_pages else 0
    br_pending = br_count - len(st.session_state.br_approved_redirects) if br_count > 0 else 0

    # Build options list
    options = []
    option_labels = []

    if broken_count > 0:
        options.append('broken_links')
        option_labels.append(f"🔗 Broken Links ({broken_pending} pending)" if broken_pending > 0 else f"🔗 Broken Links (✓ {broken_count} done)")

    if rc_count > 0:
        options.append('redirect_chains')
        option_labels.append(f"🔄 Redirect Chains ({rc_pending} pending)" if rc_pending > 0 else f"🔄 Redirect Chains (✓ {rc_count} done)")

    if iat_count > 0:
        options.append('image_alt_text')
        option_labels.append(f"🖼️ Image Alt Text ({iat_pending} pending)" if iat_pending > 0 else f"🖼️ Image Alt Text (✓ {iat_count} done)")

    if br_count > 0:
        options.append('backlink_reclaim')
        option_labels.append(f"🔙 Backlink Reclaim ({br_pending} pending)" if br_pending > 0 else f"🔙 Backlink Reclaim (✓ {br_count} done)")

    # Only show switcher if we have multiple task types
    if len(options) <= 1:
        return

    current_idx = options.index(st.session_state.current_task) if st.session_state.current_task in options else 0

    selected = st.selectbox(
        "Current Task:",
        options,
        index=current_idx,
        format_func=lambda x: option_labels[options.index(x)],
        key="task_switcher"
    )

    if selected != st.session_state.current_task:
        st.session_state.current_task = selected
        st.rerun()


def render_rc_metrics():
    """Render summary metrics for redirect chains"""
    redirects = st.session_state.rc_redirects
    decisions = st.session_state.rc_decisions
    sitewide = st.session_state.rc_sitewide
    loops = st.session_state.rc_loops
    
    total = len(redirects)
    permanent = sum(1 for r in redirects.values() if not r['is_temp_redirect'])
    temporary = total - permanent
    approved = sum(1 for d in decisions.values() if d['approved_action'])
    pending = total - approved
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Redirects", total)
    with col2:
        st.metric("301 Permanent", permanent)
    with col3:
        st.metric("302 Temporary ⚠️", temporary)
    with col4:
        st.metric("✅ Approved", approved)


def render_rc_warnings():
    """Render sitewide links and loops warnings"""
    sitewide = st.session_state.rc_sitewide
    loops = st.session_state.rc_loops
    
    if sitewide:
        with st.expander(f"⚠️ Sitewide Links Detected ({len(sitewide)}) — Manual Fix Required", expanded=False):
            st.markdown("""
            These redirect chains appear in your **header, footer, or sidebar** across your entire site.
            They should be updated manually in your theme or widget settings — fixing them in one place fixes them everywhere.
            """)
            
            for item in sitewide[:10]:  # Show first 10
                st.markdown(f"""
                <div style="background: #fef3c7; padding: 0.75rem; border-radius: 6px; margin-bottom: 0.5rem; border-left: 3px solid #f59e0b;">
                    <div style="font-family: monospace; font-size: 0.85rem; color: #92400e; word-break: break-all;">
                        {item['address'][:60]}{'...' if len(item['address']) > 60 else ''}<br>
                        → {item['final_address'][:60]}{'...' if len(item['final_address']) > 60 else ''}
                    </div>
                    <div style="font-size: 0.8rem; color: #a16207; margin-top: 0.25rem;">
                        Position: {item['position']} • Affects {item['count']} pages
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            if len(sitewide) > 10:
                st.info(f"...and {len(sitewide) - 10} more sitewide links")
    
    if loops:
        with st.expander(f"🔄 Redirect Loops Detected ({len(loops)}) — Cannot Auto-Fix", expanded=False):
            st.markdown("""
            These URLs create **infinite redirect loops**. This is typically a server configuration issue, not a content issue.
            
            💡 **Tip:** Check your `.htaccess` file or redirect plugin for conflicting rules. Contact your hosting provider if unsure.
            """)
            
            for item in loops[:10]:  # Show first 10
                st.markdown(f"""
                <div style="background: #fee2e2; padding: 0.75rem; border-radius: 6px; margin-bottom: 0.5rem; border-left: 3px solid #ef4444;">
                    <div style="font-family: monospace; font-size: 0.85rem; color: #991b1b; word-break: break-all;">
                        {item['address'][:60]}{'...' if len(item['address']) > 60 else ''}<br>
                        → {item['final_address'][:60]}{'...' if len(item['final_address']) > 60 else ''} (LOOP)
                    </div>
                    <div style="font-size: 0.8rem; color: #b91c1c; margin-top: 0.25rem;">
                        Affects {item['count']} pages
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            if len(loops) > 10:
                st.info(f"...and {len(loops) - 10} more loops")


def render_rc_spreadsheet():
    """Render the redirect chains spreadsheet view"""
    redirects = st.session_state.rc_redirects
    decisions = st.session_state.rc_decisions
    domain = st.session_state.rc_domain
    
    # User guidance
    with st.expander("📖 How to use this tool", expanded=False):
        st.markdown("""
        **Review each redirect and approve the fix:**
        
        1. **Click on a redirect row** to expand it and see your options
        2. **Review the redirect** — old URL → new URL
        3. **Approve the fix** — Screaming Frog already knows the final destination, so just approve to update your links
        4. **Export or Apply** — Once approved, export as CSV or apply directly to WordPress
        
        💡 **Tip:** Not all redirects need fixing — use your judgment. Some redirects are intentional, like affiliate links or tracking URLs that redirect to a final destination. Review before approving.
        """)
    
    # Filters
    st.markdown("**Filters:**")
    filter_cols = st.columns(4)
    with filter_cols[0]:
        st.session_state.rc_filter_301 = st.checkbox("301 Permanent", value=True, key="rc_301")
    with filter_cols[1]:
        st.session_state.rc_filter_302 = st.checkbox("302 Temporary ⚠️", value=True, key="rc_302")
    with filter_cols[2]:
        st.session_state.rc_show_pending = st.checkbox("Pending", value=True, key="rc_pending")
    with filter_cols[3]:
        st.session_state.rc_show_approved = st.checkbox("Approved", value=True, key="rc_approved")
    
    # Count pending URLs for bulk actions
    pending_keys = [k for k, d in decisions.items() if not d['approved_action']]
    pending_301 = [k for k in pending_keys if not redirects[k]['is_temp_redirect']]
    pending_302 = [k for k in pending_keys if redirects[k]['is_temp_redirect']]
    approved_count = sum(1 for d in decisions.values() if d['approved_action'])
    
    # Action bar (like Broken Links)
    st.markdown("---")
    action_cols = st.columns([2, 2, 2])
    
    with action_cols[0]:
        if len(pending_keys) > 0:
            if st.button(f"✅ Approve All ({len(pending_keys)})", type="primary", use_container_width=True):
                # Show warning if there are 302s
                if pending_302:
                    st.session_state.rc_show_approve_warning = True
                    st.rerun()
                else:
                    # No 302s, approve directly
                    for key in pending_keys:
                        st.session_state.rc_decisions[key]['approved_action'] = 'replace'
                        st.session_state.rc_decisions[key]['approved_fix'] = redirects[key]['final_address']
                    st.toast(f"✅ Approved {len(pending_keys)} redirects", icon="✅")
                    st.rerun()
        else:
            st.button("✅ Approve All", use_container_width=True, disabled=True)
    
    with action_cols[1]:
        # Placeholder for future bulk action
        pass
    
    with action_cols[2]:
        if st.button(f"📥 Export ({approved_count})", use_container_width=True, disabled=approved_count == 0):
            st.session_state.rc_scroll_to_export = True
    
    # Warning modal for 302s
    if st.session_state.get('rc_show_approve_warning'):
        st.warning(f"""
        ⚠️ **{len(pending_302)} temporary (302) redirects detected**
        
        302 redirects may revert to original URLs. Are you sure you want to approve all {len(pending_keys)} redirects?
        
        - **{len(pending_301)}** permanent (301) redirects
        - **{len(pending_302)}** temporary (302) redirects
        """)
        
        warn_cols = st.columns([1, 1, 2])
        with warn_cols[0]:
            if st.button("✅ Yes, Approve All", type="primary", use_container_width=True):
                for key in pending_keys:
                    st.session_state.rc_decisions[key]['approved_action'] = 'replace'
                    st.session_state.rc_decisions[key]['approved_fix'] = redirects[key]['final_address']
                st.session_state.rc_show_approve_warning = False
                st.toast(f"✅ Approved {len(pending_keys)} redirects", icon="✅")
                st.rerun()
        with warn_cols[1]:
            if st.button("❌ Cancel", use_container_width=True):
                st.session_state.rc_show_approve_warning = False
                st.rerun()
    
    # Apply filters
    filtered_keys = []
    for key, info in redirects.items():
        # Filter by redirect type
        if info['is_temp_redirect'] and not st.session_state.rc_filter_302:
            continue
        if not info['is_temp_redirect'] and not st.session_state.rc_filter_301:
            continue
        
        # Filter by approval status
        has_approval = bool(decisions[key]['approved_action'])
        if has_approval and not st.session_state.rc_show_approved:
            continue
        if not has_approval and not st.session_state.rc_show_pending:
            continue
        
        filtered_keys.append(key)
    
    if not filtered_keys:
        st.info("No redirects match your filters")
        return
    
    # Sort by impact (most pages affected first)
    filtered_keys.sort(key=lambda k: redirects[k]['count'], reverse=True)
    
    # Pagination
    total = len(filtered_keys)
    per_page = 10
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(st.session_state.rc_page, total_pages - 1)
    
    start = page * per_page
    end = min(start + per_page, total)
    page_keys = filtered_keys[start:end]
    
    st.markdown(f"**Showing {start+1}-{end} of {total} redirect chains**")
    
    # Column headers - matching Broken Links style
    st.markdown("""
    <div style="display: flex; background: linear-gradient(135deg, #f0fdfa 0%, #ccfbf1 100%); border: 1px solid #99f6e4; border-radius: 8px; padding: 0.5rem 0.75rem; margin: 0.5rem 0;">
        <div style="flex: 8; font-size: 0.75rem; color: #0d9488; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em;">Redirect (From → To)</div>
        <div style="flex: 1; font-size: 0.75rem; color: #0d9488; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; text-align: center;">Approve</div>
        <div style="flex: 0.8; font-size: 0.75rem; color: #0d9488; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; text-align: center;">Edit</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Render each redirect
    for key in page_keys:
        render_rc_row(key, redirects[key], decisions[key])
    
    # Pagination
    if total_pages > 1:
        pcol1, pcol2, pcol3 = st.columns([1, 2, 1])
        with pcol1:
            if st.button("← Previous", disabled=page == 0, key="rc_prev"):
                st.session_state.rc_page = page - 1
                st.rerun()
        with pcol2:
            st.markdown(f"<div style='text-align:center; padding-top: 0.5rem;'>Page {page+1} of {total_pages}</div>", unsafe_allow_html=True)
        with pcol3:
            if st.button("Next →", disabled=page >= total_pages - 1, key="rc_next"):
                st.session_state.rc_page = page + 1
                st.rerun()


def render_rc_row(key: str, info: Dict, decision: Dict):
    """Render a single redirect chain row - stacked layout with full URLs"""
    import html
    
    is_editing = st.session_state.rc_editing_url == key
    is_approved = bool(decision['approved_action'])
    
    # Full URLs for display (strip protocol for cleaner look)
    old_url_display = info['address'].replace('https://', '').replace('http://', '').replace('www.', '')
    new_url_display = info['final_address'].replace('https://', '').replace('http://', '').replace('www.', '')
    
    old_url_escaped = html.escape(info['address'])
    new_url_escaped = html.escape(info['final_address'])
    old_url_display_escaped = html.escape(old_url_display)
    new_url_display_escaped = html.escape(new_url_display)
    
    # Type badge HTML (301/302)
    badge_color = "#fef3c7" if info['is_temp_redirect'] else "#d1fae5"
    badge_text_color = "#b45309" if info['is_temp_redirect'] else "#059669"
    badge_text = "302" if info['is_temp_redirect'] else "301"
    type_badge = f"<span style='background: {badge_color}; color: {badge_text_color}; padding: 2px 6px; border-radius: 4px; font-size: 0.7rem; margin-left: 0.5rem;'>{badge_text}</span>"
    
    # Page count
    page_info = f"<span style='color: #94a3b8; font-size: 0.75rem; margin-left: 0.5rem;'>({info['count']} pages)</span>"
    
    # Build stacked row with APPROVE column
    with st.container():
        cols = st.columns([8, 1, 0.8])
        
        with cols[0]:
            # Stacked URLs with badge inline on FROM line
            if is_approved:
                st.markdown(f"""
                <div style="margin-bottom: 0.25rem;">
                    <span style="color: #64748b; font-size: 0.75rem; margin-right: 0.5rem;">FROM:</span>
                    <a href='{old_url_escaped}' target='_blank' style='color: #dc2626; text-decoration: line-through; font-size: 0.85rem; word-break: break-all;' title='Click to open'>{old_url_display_escaped}</a>
                    {type_badge}
                </div>
                <div>
                    <span style="color: #64748b; font-size: 0.75rem; margin-right: 0.5rem;">TO:</span>
                    <a href='{new_url_escaped}' target='_blank' style='color: #059669; text-decoration: none; font-size: 0.85rem; word-break: break-all;' title='Click to open'>{new_url_display_escaped}</a>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div style="margin-bottom: 0.25rem;">
                    <span style="color: #64748b; font-size: 0.75rem; margin-right: 0.5rem;">FROM:</span>
                    <a href='{old_url_escaped}' target='_blank' style='color: #dc2626; text-decoration: none; font-size: 0.85rem; word-break: break-all;' title='Click to open'>{old_url_display_escaped}</a>
                    {type_badge}{page_info}
                </div>
                <div>
                    <span style="color: #64748b; font-size: 0.75rem; margin-right: 0.5rem;">TO:</span>
                    <a href='{new_url_escaped}' target='_blank' style='color: #059669; text-decoration: none; font-size: 0.85rem; word-break: break-all;' title='Click to open'>{new_url_display_escaped}</a>
                </div>
                """, unsafe_allow_html=True)
        
        with cols[1]:
            # APPROVE column - checkbox style (like Broken Links)
            if is_approved:
                st.markdown("<span style='color: #22c55e; font-size: 1.2rem;'>✓</span>", unsafe_allow_html=True)
            else:
                # Clickable checkbox to approve directly
                if st.button("☐", key=f"rc_quick_approve_{key}", help="Approve redirect"):
                    st.session_state.rc_decisions[key]['approved_action'] = 'replace'
                    st.session_state.rc_decisions[key]['approved_fix'] = info['final_address']
                    st.toast("✅ Approved!", icon="✅")
                    st.rerun()
        
        with cols[2]:
            # Edit button
            if is_editing:
                if st.button("✕", key=f"rc_close_{key}", help="Close"):
                    st.session_state.rc_editing_url = None
                    st.rerun()
            else:
                if st.button("📝", key=f"rc_edit_{key}", help="Review details"):
                    st.session_state.rc_editing_url = key
                    st.rerun()
        
        # Compact inline edit section
        if is_editing:
            st.markdown("---")
            
            # Show warning for temp redirects
            if info['is_temp_redirect']:
                st.warning("⚠️ **302 Temporary** redirect - destination may revert. Only approve if confident it's stable.")
            
            # Affected pages in compact expander
            with st.expander(f"📄 {len(info['sources'])} affected pages", expanded=False):
                for source in info['sources'][:10]:
                    short_source = source.replace('https://', '').replace('http://', '')[:70]
                    st.markdown(f"- `{short_source}`")
                if len(info['sources']) > 10:
                    st.caption(f"...and {len(info['sources']) - 10} more")
            
            # Action buttons - compact row
            btn_cols = st.columns([1, 1, 2])
            with btn_cols[0]:
                if not is_approved:
                    if st.button("✅ Approve", key=f"rc_approve_{key}", type="primary", use_container_width=True):
                        st.session_state.rc_decisions[key]['approved_action'] = 'replace'
                        st.session_state.rc_decisions[key]['approved_fix'] = info['final_address']
                        st.session_state.rc_editing_url = None
                        st.toast("✅ Approved!", icon="✅")
                        st.rerun()
            
            with btn_cols[1]:
                if is_approved:
                    if st.button("↩️ Reset", key=f"rc_reset_{key}", use_container_width=True):
                        st.session_state.rc_decisions[key]['approved_action'] = ''
                        st.session_state.rc_decisions[key]['approved_fix'] = ''
                        st.session_state.rc_editing_url = None
                        st.toast("↩️ Reset", icon="↩️")
                        st.rerun()
            
            st.markdown("---")


def render_rc_export_section():
    """Render export section for redirect chains"""
    decisions = st.session_state.rc_decisions
    redirects = st.session_state.rc_redirects
    
    approved = [(k, d) for k, d in decisions.items() if d['approved_action']]
    has_approved = len(approved) > 0
    
    st.markdown('<p class="section-header">Export & Apply</p>', unsafe_allow_html=True)
    
    # Status message based on approved count
    if has_approved:
        # Count total pages affected
        total_pages = sum(len(redirects[k]['sources']) for k, d in approved)
        st.markdown(f"✅ **{len(approved)} redirects approved** affecting {total_pages} pages")
    else:
        st.markdown("No redirects approved yet. Approve fixes above to enable export.")
    
    # Export buttons
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📥 Export CSV", use_container_width=True, disabled=not has_approved, key="rc_export_csv"):
            export_data = create_rc_export_data()
            
            track_event("export", {
                "format": "csv",
                "type": "redirect_chains",
                "total_fixes": len(export_data),
            })
            
            csv_data = pd.DataFrame(export_data).to_csv(index=False)
            st.download_button(
                "Download CSV",
                data=csv_data,
                file_name=f"redirect_chain_fixes_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
                key="rc_download_csv"
            )
    
    with col2:
        if st.button("📥 Export JSON", use_container_width=True, disabled=not has_approved, key="rc_export_json"):
            export_data = create_rc_export_data()
            
            track_event("export", {
                "format": "json",
                "type": "redirect_chains",
                "total_fixes": len(export_data),
            })
            
            json_data = json.dumps(export_data, indent=2)
            st.download_button(
                "Download JSON",
                data=json_data,
                file_name=f"redirect_chain_fixes_{datetime.now().strftime('%Y%m%d')}.json",
                mime="application/json",
                use_container_width=True,
                key="rc_download_json"
            )
    
    # WordPress apply section
    if not st.session_state.wp_connected:
        return
    
    st.markdown("---")
    st.markdown("**Apply directly to WordPress:**")
    
    if not has_approved:
        st.info("Approve some redirects above first")
        return
    
    # Count how many source pages need fixes
    source_pages_to_fix = set()
    for key, decision in approved:
        for source in redirects[key]['sources']:
            source_pages_to_fix.add(source)
    
    pages_to_fix = len(source_pages_to_fix)
    selected_mode = st.session_state.get('selected_mode', 'quick_start')
    
    if selected_mode == 'quick_start' and pages_to_fix <= AGENT_MODE_LIMIT:
        st.info(f"🤖 **Quick Start Mode:** {pages_to_fix} pages to update")
        
        if not st.session_state.post_id_check_done:
            st.markdown("Before applying, let's verify your site supports automatic Post ID discovery.")
            if st.button("🔍 Run Compatibility Check", use_container_width=True, key="rc_check"):
                run_post_id_sample_check(list(source_pages_to_fix)[:3])
        else:
            if st.session_state.post_id_check_passed:
                st.success("✅ Compatibility check passed!")
                if st.button("🚀 Apply Redirect Fixes to WordPress", type="primary", use_container_width=True, key="rc_apply"):
                    run_rc_agent_fixes(approved, source_pages_to_fix)
            else:
                st.error("❌ Automatic Post ID lookup failed.")
                st.info("💡 Try setting up Full Mode for more reliable Post ID handling.")
    else:
        st.warning(f"⚠️ {pages_to_fix} pages exceeds Quick Start Mode limit ({AGENT_MODE_LIMIT}). Use Full Mode or export CSV for manual updates.")


def run_rc_agent_fixes(approved: List, source_pages_to_fix: set):
    """Run the agent to apply redirect chain fixes"""
    client = st.session_state.wp_client
    redirects = st.session_state.rc_redirects
    
    # Build list of all fixes to apply
    fixes_to_apply = []
    for key, decision in approved:
        info = redirects[key]
        for source in info['sources']:
            if source in source_pages_to_fix:
                fixes_to_apply.append({
                    'source_url': source,
                    'old_url': info['address'],
                    'new_url': info['final_address'],
                    'post_id': st.session_state.post_id_cache.get(source)
                })
    
    total_fixes = len(fixes_to_apply)
    results = []
    
    st.markdown("### 🔄 Applying Redirect Fixes")
    
    progress_bar = st.progress(0)
    status_container = st.container()
    
    with status_container:
        for i, fix in enumerate(fixes_to_apply):
            progress_bar.progress((i + 1) / total_fixes)
            
            with st.status(f"Fix {i+1} of {total_fixes}", expanded=True) as status:
                # Step 1: Get Post ID
                post_id = fix['post_id']
                
                if not post_id:
                    st.write(f"🔍 Finding Post ID for: `{fix['source_url'][:50]}...`")
                    post_id = client.find_post_id_by_url(fix['source_url'])
                    
                    if post_id:
                        st.write(f"   ✅ Found Post ID: {post_id}")
                        st.session_state.post_id_cache[fix['source_url']] = post_id
                    else:
                        # Check if it's a category/archive/tag page
                        source_lower = fix['source_url'].lower()
                        is_archive = any(x in source_lower for x in ['/category/', '/tag/', '/author/', '/page/', '/archive/'])
                        
                        if is_archive:
                            st.write(f"   ℹ️ This is an archive/category page (dynamically generated)")
                            st.write(f"   💡 Fix the links on individual posts — archive pages will update automatically")
                            msg = 'Archive/category page - fix individual posts instead'
                        else:
                            st.write(f"   ❌ Post ID not found — skipping")
                            msg = 'Post ID not found'
                        
                        results.append({
                            'source_url': fix['source_url'],
                            'old_url': fix['old_url'],
                            'new_url': fix['new_url'],
                            'status': 'skipped',
                            'message': msg
                        })
                        status.update(label=f"⏭️ Skipped", state="error")
                        continue
                else:
                    st.write(f"📄 Post ID: {post_id}")
                
                # Step 2: Apply fix (replace old URL with new URL)
                st.write(f"🔧 Replacing URL...")
                st.write(f"   Old: `{fix['old_url'][:40]}...`")
                st.write(f"   New: `{fix['new_url'][:40]}...`")
                
                try:
                    result = client.replace_link(post_id, fix['old_url'], fix['new_url'], dry_run=False)
                    
                    if result['success']:
                        st.write(f"   ✅ {result['message']}")
                        results.append({
                            'source_url': fix['source_url'],
                            'old_url': fix['old_url'],
                            'new_url': fix['new_url'],
                            'status': 'success',
                            'message': result['message']
                        })
                        status.update(label=f"✅ Success", state="complete")
                    else:
                        # Provide helpful context for "URL not found" errors
                        if 'not found' in result['message'].lower():
                            st.write(f"   ❌ {result['message']}")
                            st.write(f"   💡 Common reasons: link may be in a widget, shortcode, custom field, or theme template")
                        else:
                            st.write(f"   ❌ {result['message']}")
                        
                        results.append({
                            'source_url': fix['source_url'],
                            'old_url': fix['old_url'],
                            'new_url': fix['new_url'],
                            'status': 'failed',
                            'message': result['message']
                        })
                        status.update(label=f"❌ Failed", state="error")
                        
                except Exception as e:
                    st.write(f"   ❌ Error: {str(e)}")
                    results.append({
                        'source_url': fix['source_url'],
                        'old_url': fix['old_url'],
                        'new_url': fix['new_url'],
                        'status': 'failed',
                        'message': str(e)
                    })
                    status.update(label=f"❌ Error", state="error")
            
            time.sleep(0.3)
    
    # Store results
    st.session_state.wp_execute_results = results
    
    # Summary
    success = sum(1 for r in results if r['status'] == 'success')
    skipped = sum(1 for r in results if r['status'] == 'skipped')
    failed = sum(1 for r in results if r['status'] == 'failed')
    
    track_event("wordpress_apply", {
        "type": "redirect_chains",
        "total_fixes": len(results),
        "success": success,
        "skipped": skipped,
        "failed": failed
    })
    
    st.markdown("---")
    st.markdown("### 📊 Execution Complete")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("✅ Success", success)
    with col2:
        st.metric("⏭️ Skipped", skipped)
    with col3:
        st.metric("❌ Failed", failed)


# =============================================================================
# IMAGE ALT TEXT UI COMPONENTS
# =============================================================================

def render_iat_metrics():
    """Render summary metrics for image alt text"""
    images = st.session_state.iat_images
    decisions = st.session_state.iat_decisions
    excluded = st.session_state.iat_excluded_count
    
    total = len(images)
    approved = sum(1 for d in decisions.values() if d['approved_action'])
    pending = total - approved
    
    # Count by alt status
    missing = sum(1 for img in images.values() if img['alt_status'] == 'missing')
    filename = sum(1 for img in images.values() if img['alt_status'] == 'filename')
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Images", total)
    with col2:
        st.metric("Missing Alt", missing)
    with col3:
        st.metric("Filename Alt", filename)
    with col4:
        st.metric("Pending Review", pending)
    
    if excluded > 0:
        st.caption(f"ℹ️ {excluded} images were filtered out (logos, icons, non-content pages, or already have good alt text)")


def render_iat_spreadsheet():
    """Render the image alt text spreadsheet view"""
    images = st.session_state.iat_images
    decisions = st.session_state.iat_decisions
    
    if not images:
        st.info("No images to display")
        return
    
    # Filters
    col1, col2 = st.columns(2)
    with col1:
        show_pending = st.checkbox("Show Pending", value=st.session_state.iat_show_pending, key="iat_filter_pending")
        st.session_state.iat_show_pending = show_pending
    with col2:
        show_approved = st.checkbox("Show Approved", value=st.session_state.iat_show_approved, key="iat_filter_approved")
        st.session_state.iat_show_approved = show_approved
    
    # Filter images
    filtered_keys = []
    for key in images:
        has_approval = bool(decisions[key]['approved_action'])
        if has_approval and show_approved:
            filtered_keys.append(key)
        elif not has_approval and show_pending:
            filtered_keys.append(key)
    
    # Sort by impact (pages affected)
    filtered_keys.sort(key=lambda k: images[k]['count'], reverse=True)
    
    # Pagination
    page = st.session_state.iat_page
    per_page = 10
    total_pages = max(1, (len(filtered_keys) + per_page - 1) // per_page)
    
    if page >= total_pages:
        page = total_pages - 1
        st.session_state.iat_page = page
    
    start_idx = page * per_page
    end_idx = min(start_idx + per_page, len(filtered_keys))
    page_keys = filtered_keys[start_idx:end_idx]
    
    st.markdown(f"**Showing {start_idx + 1}-{end_idx} of {len(filtered_keys)} images**")
    
    # Render each image row
    for i, key in enumerate(page_keys):
        is_first_row = (i == 0 and page == 0 and not any(d['approved_action'] for d in decisions.values()))
        render_iat_row(key, images[key], decisions[key], is_first_row)
    
    # Pagination controls
    if total_pages > 1:
        pcol1, pcol2, pcol3 = st.columns([1, 2, 1])
        with pcol1:
            if st.button("← Previous", disabled=page == 0, key="iat_prev"):
                st.session_state.iat_page = page - 1
                st.rerun()
        with pcol2:
            st.markdown(f"<div style='text-align:center; padding-top: 0.5rem;'>Page {page+1} of {total_pages}</div>", unsafe_allow_html=True)
        with pcol3:
            if st.button("Next →", disabled=page >= total_pages - 1, key="iat_next"):
                st.session_state.iat_page = page + 1
                st.rerun()


def render_iat_row(img_url: str, info: Dict, decision: Dict, is_first_row: bool = False):
    """Render a single image alt text row - simplified view"""
    import html
    from urllib.parse import unquote
    
    # Approval status
    if decision['approved_action'] == 'ignore':
        status_text = "⏭️ Ignored"
        status_color = "#64748b"
    elif decision['approved_action'] == 'replace':
        status_text = "✅ Fixed"
        status_color = "#059669"
    else:
        status_text = "⏳ Pending"
        status_color = "#b45309"
    
    # Image filename - decode and escape
    img_url_decoded = unquote(img_url)
    img_filename = img_url_decoded.split('/')[-1][:50]
    img_filename_escaped = html.escape(img_filename)
    img_url_escaped = html.escape(img_url)
    
    # Simple row with thumbnail, filename, status, page count
    col1, col2, col3, col4 = st.columns([1, 3, 1, 1])
    
    with col1:
        # Thumbnail
        try:
            st.image(img_url, width=60)
        except:
            st.caption("No preview")
    
    with col2:
        st.markdown(f"**{img_filename_escaped}**")
        st.caption(f"Affects {info['count']} page(s)")
    
    with col3:
        st.markdown(f"<span style='color: {status_color}; font-weight: 500;'>{status_text}</span>", unsafe_allow_html=True)
    
    with col4:
        is_editing = st.session_state.iat_editing_url == img_url
        if is_editing:
            if st.button("✕ Close", key=f"iat_close_{img_url}", use_container_width=True):
                st.session_state.iat_editing_url = None
                st.rerun()
        else:
            if st.button("Set Fix", key=f"iat_edit_{img_url}", type="primary", use_container_width=True):
                st.session_state.iat_editing_url = img_url
                st.rerun()
    
    # Show inline edit form if this image is being edited
    if st.session_state.iat_editing_url == img_url:
        render_iat_edit_form(img_url, info, decision)


def render_iat_edit_form(img_url: str, info: Dict, decision: Dict):
    """Render the inline edit form for image alt text"""
    
    # Use a Streamlit container with custom styling via CSS class
    with st.container():
        # Show image thumbnail
        st.markdown("**Image Preview:**")
        try:
            st.image(img_url, width=300)
        except:
            st.caption("(Could not load image preview)")
        
        st.markdown("**Choose a fix for this image:**")
        
        # Fix options
        fix_options = ["✏️ Use Manual", "🤖 Use AI", "⏭️ Ignore"]
        
        # Determine default selection
        if decision['approved_action'] == 'replace':
            default_idx = 0
        elif decision['approved_action'] == 'ignore':
            default_idx = 2
        else:
            default_idx = 0
        
        selected_fix = st.radio(
            "Fix type",
            fix_options,
            index=default_idx,
            horizontal=True,
            key=f"iat_fix_type_{img_url}",
            label_visibility='collapsed'
        )
    
    # Reduced spacing divider (half the height of st.markdown("---"))
    st.markdown("<div style='border-top: 1px solid #e2e8f0; margin: 0.5rem 0;'></div>", unsafe_allow_html=True)
    
    if selected_fix == "✏️ Use Manual":
        st.markdown("**What this does:** Sets the alt text to the value you provide. Use when you know exactly what the image shows.")
        
        # Text input - get the current value
        current_value = decision['manual_fix'] or decision['approved_fix'] or ''
        
        # Label with character count on same row using columns
        col_label, col_chars = st.columns([1, 1])
        with col_label:
            st.markdown("**Enter alt text:**")
        with col_chars:
            # Show character count based on current stored value
            char_count = len(current_value)
            if char_count > 0:
                if char_count < 10:
                    st.markdown(f"<div style='text-align: right; color: #d97706; font-size: 0.85rem;'>⚠️ {char_count} chars (try to be more descriptive)</div>", unsafe_allow_html=True)
                elif char_count <= 125:
                    st.markdown(f"<div style='text-align: right; color: #059669; font-size: 0.85rem;'>✅ {char_count} chars (good length)</div>", unsafe_allow_html=True)
                else:
                    st.markdown(f"<div style='text-align: right; color: #d97706; font-size: 0.85rem;'>⚠️ {char_count} chars (consider shortening)</div>", unsafe_allow_html=True)
        
        # Text input
        manual_alt = st.text_input(
            "Enter alt text:",
            value=current_value,
            placeholder="Descriptive alt text for this image",
            key=f"iat_manual_{img_url}",
            max_chars=200,
            label_visibility='collapsed'
        )
        
        if manual_alt != decision['manual_fix']:
            st.session_state.iat_decisions[img_url]['manual_fix'] = manual_alt
        
        # Always show Save button - enabled state, check for content on click
        if st.button("💾 Save Selection", key=f"iat_save_manual_{img_url}", type="primary", use_container_width=True):
            if manual_alt:
                st.session_state.iat_decisions[img_url]['approved_action'] = 'replace'
                st.session_state.iat_decisions[img_url]['approved_fix'] = manual_alt
                st.session_state.iat_editing_url = None
                st.toast("✅ Saved: Replace Alt Text", icon="✅")
                time.sleep(0.3)
                st.rerun()
            else:
                st.warning("Please enter alt text before saving")
    
    elif selected_fix == "🤖 Use AI":
        st.markdown("**What this does:** AI analyzes the actual image and generates descriptive alt text. Use when you want help describing the image content.")
        
        # Check for API key
        user_has_key = bool(st.session_state.anthropic_key)
        
        # Show existing AI suggestion if available
        if decision['ai_suggestion']:
            st.markdown("<div style='border-top: 1px solid #e2e8f0; margin: 0.5rem 0;'></div>", unsafe_allow_html=True)
            st.markdown("**AI Recommendation:**")
            st.success(f'"{decision["ai_suggestion"]}"')
            
            if decision['ai_notes']:
                st.markdown(f"**Why:** {decision['ai_notes']}")
            
            if st.button("💾 Accept AI Suggestion", key=f"iat_save_ai_{img_url}", type="primary", use_container_width=True):
                st.session_state.iat_decisions[img_url]['approved_action'] = 'replace'
                st.session_state.iat_decisions[img_url]['approved_fix'] = decision['ai_suggestion']
                st.session_state.iat_editing_url = None
                st.toast("✅ Saved: Replace Alt Text", icon="✅")
                time.sleep(0.3)
                st.rerun()
        else:
            if user_has_key:
                st.success("✅ Using your API key")
                if st.button("🔍 Get AI Suggestion", key=f"iat_ai_{img_url}", use_container_width=True):
                    with st.spinner("AI is analyzing the image..."):
                        domain = st.session_state.iat_domain or ''
                        result = get_ai_alt_text_suggestion(img_url, info, domain, st.session_state.anthropic_key)
                        st.session_state.iat_decisions[img_url]['ai_suggestion'] = result['alt_text']
                        st.session_state.iat_decisions[img_url]['ai_notes'] = result['notes']
                        st.rerun()
            else:
                st.info("🔑 Enter your Claude API key for AI-powered alt text suggestions")
                
                api_key_input = st.text_input(
                    "Your Claude API Key:",
                    type="password",
                    placeholder="sk-ant-...",
                    key=f"iat_api_key_{img_url}",
                )
                if api_key_input:
                    st.session_state.anthropic_key = api_key_input
                    st.success("✅ API key saved!")
                    st.rerun()
                
                st.caption("Get your key at [console.anthropic.com](https://console.anthropic.com) • Keys are stored in your browser session only")
    
    elif selected_fix == "⏭️ Ignore":
        st.markdown("**What this does:** Marks this image as reviewed but takes no action. Use for decorative images, images with acceptable alt text, or ones you'll handle manually.")
        
        if st.button("💾 Save Selection", key=f"iat_save_ignore_{img_url}", type="primary", use_container_width=True):
            st.session_state.iat_decisions[img_url]['approved_action'] = 'ignore'
            st.session_state.iat_decisions[img_url]['approved_fix'] = ''
            st.session_state.iat_editing_url = None
            st.toast("✅ Saved: Ignored", icon="✅")
            time.sleep(0.3)
            st.rerun()


def create_iat_export_data() -> List[Dict]:
    """Create export data from approved image alt text decisions"""
    decisions = st.session_state.iat_decisions
    images = st.session_state.iat_images
    
    export = []
    
    for img_url, decision in decisions.items():
        if not decision['approved_action'] or decision['approved_action'] == 'ignore':
            continue
        
        info = images[img_url]
        
        for source in info['sources']:
            export.append({
                'source_url': source,
                'image_url': img_url,
                'current_alt': info['current_alt'],
                'new_alt': decision['approved_fix'],
                'action': 'replace',
            })
    
    return export


def render_iat_export_section():
    """Render the export section for image alt text"""
    decisions = st.session_state.iat_decisions
    images = st.session_state.iat_images
    
    # Count approved
    approved = [(k, v) for k, v in decisions.items() if v['approved_action'] == 'replace']
    ignored = sum(1 for v in decisions.values() if v['approved_action'] == 'ignore')
    pending = len(decisions) - len(approved) - ignored
    
    has_approved = len(approved) > 0
    
    st.markdown(f"""
    **Summary:** {len(approved)} approved • {ignored} ignored • {pending} pending
    """)
    
    if not has_approved:
        st.info("Set alt text fixes above to enable export")
        return
    
    # Export buttons
    export_data = create_iat_export_data()
    
    col1, col2 = st.columns(2)
    
    with col1:
        csv_output = pd.DataFrame(export_data).to_csv(index=False)
        st.download_button(
            label="📥 Download CSV",
            data=csv_output,
            file_name=f"image_alt_fixes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            use_container_width=True
        )
    
    with col2:
        json_output = json.dumps(export_data, indent=2)
        st.download_button(
            label="📥 Download JSON",
            data=json_output,
            file_name=f"image_alt_fixes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json",
            use_container_width=True
        )
    
    # WordPress apply section
    if not st.session_state.wp_connected:
        st.markdown("---")
        st.info("💡 Connect to WordPress above to apply fixes directly to your site")
        return
    
    st.markdown("---")
    st.markdown("**Apply directly to WordPress:**")
    
    if not has_approved:
        st.info("Approve some alt text fixes above first")
        return
    
    # Check if we have Post IDs
    has_post_ids = st.session_state.has_post_ids
    post_id_count = len(st.session_state.post_id_cache)
    
    # Count how many source pages need fixes
    source_pages_to_fix = set()
    for key, decision in approved:
        for source in images[key]['sources']:
            source_pages_to_fix.add(source)
    
    pages_to_fix = len(source_pages_to_fix)
    
    # Count how many pages have Post IDs mapped
    pages_with_post_ids = sum(1 for url in source_pages_to_fix if url in st.session_state.post_id_cache)
    
    if has_post_ids and pages_with_post_ids > 0:
        # Full Mode with Post IDs
        st.success(f"🚀 **Full Mode:** {pages_with_post_ids} of {pages_to_fix} pages have Post IDs mapped")
        
        if pages_with_post_ids < pages_to_fix:
            st.warning(f"⚠️ {pages_to_fix - pages_with_post_ids} pages don't have Post IDs and will be skipped")
        
        if st.button("🚀 Apply Alt Text Fixes to WordPress", type="primary", use_container_width=True, key="iat_apply"):
            run_iat_agent_fixes(approved, source_pages_to_fix)
    else:
        # No Post IDs - explain what's needed
        st.warning(f"""
        **Post IDs Required**
        
        To apply fixes to WordPress, you need to upload a Post ID mapping first.
        
        **{pages_to_fix} pages** need updates, but we don't have Post IDs for them.
        
        👆 Upload your Post IDs CSV in the upload section above, then return here.
        """)


def run_iat_agent_fixes(approved: List, source_pages_to_fix: set):
    """Run the agent to apply image alt text fixes"""
    client = st.session_state.wp_client
    images = st.session_state.iat_images
    
    # Build list of all fixes to apply
    fixes_to_apply = []
    for img_url, decision in approved:
        info = images[img_url]
        for source in info['sources']:
            if source in source_pages_to_fix:
                fixes_to_apply.append({
                    'source_url': source,
                    'image_url': img_url,
                    'new_alt': decision['approved_fix'],
                    'post_id': st.session_state.post_id_cache.get(source)
                })
    
    total_fixes = len(fixes_to_apply)
    results = []
    skipped_for_retry = []
    
    # Helper function to truncate with ellipsis if needed
    def truncate(text, max_len):
        if len(text) <= max_len:
            return text
        return text[:max_len] + "..."
    
    with st.status(f"Applying {total_fixes} alt text fixes...", expanded=True) as main_status:
        for i, fix in enumerate(fixes_to_apply):
            with st.status(f"Fix {i+1}/{total_fixes}", expanded=False) as status:
                # Clickable page link (truncate display but full link)
                page_url = fix['source_url']
                page_display = truncate(page_url, 100)
                st.markdown(f"📄 Page: [{page_display}]({page_url})")
                
                # Step 1: Get Post ID
                post_id = fix.get('post_id')
                if not post_id:
                    st.write("🔍 Looking up Post ID...")
                    post_id = client.find_post_id_by_url(fix['source_url'])
                    
                    if post_id:
                        st.write(f"✅ Found Post ID: {post_id}")
                        st.session_state.post_id_cache[fix['source_url']] = post_id
                    else:
                        st.write("❌ Could not find Post ID")
                        results.append({
                            'source_url': fix['source_url'],
                            'image_url': fix['image_url'],
                            'status': 'skipped',
                            'message': 'Post ID not found'
                        })
                        skipped_for_retry.append(fix)
                        status.update(label=f"⏭️ Skipped", state="error")
                        continue
                else:
                    st.write(f"📄 Post ID: {post_id}")
                
                # Step 2: Apply fix (update alt text)
                st.write(f"🔧 Updating alt text...")
                
                # Full image filename (truncate at 80 chars)
                img_filename = fix['image_url'].split('/')[-1]
                img_display = truncate(img_filename, 80)
                st.write(f"   Image: `{img_display}`")
                
                # Full alt text (truncate at 150 chars)
                alt_display = truncate(fix['new_alt'], 150)
                st.write(f"   New alt: `{alt_display}`")
                
                try:
                    result = client.update_alt_text(post_id, fix['image_url'], fix['new_alt'], dry_run=False)
                    
                    if result['success']:
                        st.write(f"   ✅ {result['message']}")
                        results.append({
                            'source_url': fix['source_url'],
                            'image_url': fix['image_url'],
                            'new_alt': fix['new_alt'],
                            'status': 'success',
                            'message': result['message']
                        })
                        status.update(label=f"✅ Success", state="complete")
                    else:
                        if 'not found' in result['message'].lower():
                            st.write(f"   ❌ {result['message']}")
                            st.write(f"   💡 Common reasons: image may be in a widget, shortcode, or theme template")
                        else:
                            st.write(f"   ❌ {result['message']}")
                        
                        results.append({
                            'source_url': fix['source_url'],
                            'image_url': fix['image_url'],
                            'new_alt': fix['new_alt'],
                            'status': 'failed',
                            'message': result['message']
                        })
                        status.update(label=f"❌ Failed", state="error")
                        
                except Exception as e:
                    st.write(f"   ❌ Error: {str(e)}")
                    results.append({
                        'source_url': fix['source_url'],
                        'image_url': fix['image_url'],
                        'new_alt': fix['new_alt'],
                        'status': 'failed',
                        'message': str(e)
                    })
                    status.update(label=f"❌ Error", state="error")
            
            time.sleep(0.3)
    
    # Store results
    st.session_state.wp_execute_results = results
    
    # Summary
    success = sum(1 for r in results if r['status'] == 'success')
    skipped = sum(1 for r in results if r['status'] == 'skipped')
    failed = sum(1 for r in results if r['status'] == 'failed')
    
    track_event("wordpress_apply", {
        "type": "image_alt_text",
        "total_fixes": len(results),
        "success": success,
        "skipped": skipped,
        "failed": failed
    })
    
    st.markdown("---")
    st.markdown("### 📊 Execution Complete")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("✅ Success", success)
    with col2:
        st.metric("⏭️ Skipped", skipped)
    with col3:
        st.metric("❌ Failed", failed)


def render_about():
    """Render the About section"""
    st.markdown('<p class="section-header" id="about">About</p>', unsafe_allow_html=True)
    
    st.markdown("""
    **Screaming Fixes** is a free, open-source tool built for SEO professionals who are tired of manually 
    updating broken links, redirect chains, and image alt text one by one.
    
    *We are not affiliated with, sponsored by, or endorsed by [Screaming Frog SEO Spider](https://www.screamingfrog.co.uk/seo-spider/). 
    We just love using their tool and wanted to help SEOs get even more value from it.*
    
    **What this tool does:**
    - **Broken Links:** Groups thousands of link references by unique broken URL, with optional AI suggestions
    - **Redirect Chains:** Bulk updates outdated URLs to their final destinations
    - **Image Alt Text:** Identifies images with missing or non-descriptive alt text, with AI-powered suggestions
    - Requires **human approval** for all fixes before they're applied
    - **Fixes and publishes updates** directly to your live WordPress website
    - Exports to CSV/JSON for use with WordPress or any CMS
    
    **Key insight:** Your Screaming Frog export might have 7,000+ link references, but only 60-100 unique URLs to fix. 
    Fix each unique URL once, and you've fixed them all.
    
    Your data never leaves your browser session. [View the source code](https://github.com/backofnapkin/screaming-fixes) to verify.
    """)
    
    st.markdown("""
    ---
    **⚠️ Disclaimer:** This tool is provided "as-is" without warranty of any kind. You are solely responsible for 
    reviewing, approving, and applying any changes to your website. Screaming Frog data may occasionally contain 
    errors or false positives. Always back up your site before making bulk changes. The creators of this tool 
    are not liable for any damages or issues resulting from its use.
    """)


def render_instructions():
    """Render the Instructions section"""
    st.markdown('<p class="section-header" id="instructions">Instructions</p>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        with st.expander("📋 How to Export Broken Links", expanded=True):
            st.markdown("""
            1. Run your crawl in Screaming Frog
            2. Go to **Bulk Export** menu
            3. Select **Response Codes → Client Error (4xx) → Inlinks**
            4. Save the CSV file
            5. Upload it here
            
            This exports all internal pages that link to broken URLs, 
            including the anchor text used.
            """)
        
        with st.expander("🔄 How to Export Redirect Chains"):
            st.markdown("""
            1. Run your crawl in Screaming Frog
            2. Go to **Reports** menu
            3. Select **Redirects → All Redirects**
            4. Save the CSV file
            5. Upload it here
            
            This exports all redirect chains found during the crawl,
            along with the final destination URLs.
            """)
        
        with st.expander("🤖 Getting AI Suggestions (Broken Links only)"):
            st.markdown("""
            1. Get an API key from [console.anthropic.com](https://console.anthropic.com)
            2. Paste it in the "AI Suggestions" section
            3. Click "Get AI Suggestion" on any broken URL
            4. Claude will search the web for replacements
            5. **Review and approve** before applying
            
            *Cost: ~$0.01-0.05 per suggestion with web search*
            
            **Note:** Redirect chains don't need AI — the replacement URL is already known!
            """)
    
    with col2:
        with st.expander("🔐 WordPress Application Passwords"):
            st.markdown("""
            Application Passwords let external apps access your site securely:
            
            1. Go to **WordPress Admin → Users → Profile**
            2. Scroll to **Application Passwords**
            3. Enter name: `Screaming Fixes`
            4. Click **Add New Application Password**
            5. Copy the password (spaces are OK)
            
            *Requires WordPress 5.6+ or the Application Passwords plugin*
            """)
        
        with st.expander("✅ Approving & Applying Fixes"):
            st.markdown("""
            For each broken URL, you can:
            
            - **🗑️ Remove** — Delete the link, keep the anchor text
            - **✓ Use Manual** — Use the URL you typed
            - **✓ Use AI** — Accept Claude's suggestion
            
            Then either:
            - **Export CSV/JSON** for manual updates
            - **Apply to WordPress** directly via REST API
            
            Nothing is applied until you approve it.
            """)


def render_footer():
    """Render footer"""
    st.markdown("---")
    st.markdown("""
    <div class="footer">
        <div style="display: flex; justify-content: center; gap: 8px; margin-bottom: 1rem;">
            <a href="https://github.com/backofnapkin/screaming-fixes" target="_blank">
                <img src="https://img.shields.io/badge/GitHub-View_Source-134e4a?style=flat-square&logo=github" alt="GitHub">
            </a>
            <a href="https://streamlit.io" target="_blank">
                <img src="https://img.shields.io/badge/Powered_by-Streamlit-14b8a6?style=flat-square" alt="Streamlit">
            </a>
        </div>
        Built with ❤️ for the SEO community
        <div style="margin-top: 1rem; font-size: 0.75rem; color: #94a3b8;">
            <strong>Privacy:</strong> Your credentials and file contents are processed in your browser session only and are never stored. 
            We collect anonymous usage analytics (feature usage, error rates) to improve the tool. No personal data or URLs are tracked.
        </div>
    </div>
    """, unsafe_allow_html=True)


# =============================================================================
# MAIN
# =============================================================================

def render_progress_indicator():
    """Render workflow progress indicator"""
    decisions = st.session_state.decisions
    
    # Determine current state
    has_upload = st.session_state.df is not None
    has_reviewed = any(d['approved_action'] for d in decisions.values()) if decisions else False
    all_reviewed = all(d['approved_action'] for d in decisions.values()) if decisions else False
    approved_count = sum(1 for d in decisions.values() if d['approved_action'] and d['approved_action'] != 'ignore')
    has_exports = approved_count > 0
    
    # Step states: 'complete', 'active', 'pending'
    step1 = 'complete' if has_upload else 'active'
    step2 = 'complete' if all_reviewed else ('active' if has_upload else 'pending')
    step3 = 'complete' if has_exports else ('active' if has_reviewed else 'pending')
    step4 = 'active' if has_exports else 'pending'
    
    def get_circle_style(state):
        if state == 'complete':
            return 'background: linear-gradient(135deg, #0d9488, #0891b2); color: white;'
        elif state == 'active':
            return 'background: linear-gradient(135deg, #ccfbf1, #a5f3fc); color: #0d9488; border: 2px solid #0d9488;'
        else:
            return 'background: #e2e8f0; color: #94a3b8;'
    
    def get_icon(state, num):
        return '✓' if state == 'complete' else num
    
    def get_connector_style(prev_state):
        if prev_state == 'complete':
            return 'background: linear-gradient(90deg, #0d9488, #0891b2);'
        else:
            return 'background: #e2e8f0;'
    
    def get_label_style(state):
        if state == 'complete':
            return 'color: #0d9488; font-weight: 600;'
        elif state == 'active':
            return 'color: #0d9488; font-weight: 600;'
        else:
            return 'color: #94a3b8;'
    
    st.markdown(f"""
    <div style="display: flex; align-items: center; justify-content: center; margin: 0.5rem 0 1.5rem 0;">
        <!-- Step 1 -->
        <div style="display: flex; flex-direction: column; align-items: center; min-width: 70px;">
            <div style="width: 32px; height: 32px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 0.85rem; {get_circle_style(step1)}">{get_icon(step1, '1')}</div>
            <div style="font-size: 0.75rem; margin-top: 0.35rem; {get_label_style(step1)}">Upload</div>
        </div>
        <!-- Connector 1-2 -->
        <div style="width: 40px; height: 3px; {get_connector_style(step1)} margin: 0 -5px; margin-bottom: 1.2rem;"></div>
        <!-- Step 2 -->
        <div style="display: flex; flex-direction: column; align-items: center; min-width: 70px;">
            <div style="width: 32px; height: 32px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 0.85rem; {get_circle_style(step2)}">{get_icon(step2, '2')}</div>
            <div style="font-size: 0.75rem; margin-top: 0.35rem; {get_label_style(step2)}">Review</div>
        </div>
        <!-- Connector 2-3 -->
        <div style="width: 40px; height: 3px; {get_connector_style(step2)} margin: 0 -5px; margin-bottom: 1.2rem;"></div>
        <!-- Step 3 -->
        <div style="display: flex; flex-direction: column; align-items: center; min-width: 70px;">
            <div style="width: 32px; height: 32px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 0.85rem; {get_circle_style(step3)}">{get_icon(step3, '3')}</div>
            <div style="font-size: 0.75rem; margin-top: 0.35rem; {get_label_style(step3)}">Approve</div>
        </div>
        <!-- Connector 3-4 -->
        <div style="width: 40px; height: 3px; {get_connector_style(step3)} margin: 0 -5px; margin-bottom: 1.2rem;"></div>
        <!-- Step 4 -->
        <div style="display: flex; flex-direction: column; align-items: center; min-width: 70px;">
            <div style="width: 32px; height: 32px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 0.85rem; {get_circle_style(step4)}">{get_icon(step4, '4')}</div>
            <div style="font-size: 0.75rem; margin-top: 0.35rem; {get_label_style(step4)}">Apply</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def main():
    init_session_state()
    render_nav()
    render_header()
    render_feature_cards()
    
    # Show integrations panel if requested or not all connected
    if st.session_state.show_integrations:
        render_integrations_panel(process_post_id_upload)
    
    # Always show upload section
    render_upload_section()
    
    render_progress_indicator()
    
    # Initialize backlink reclaim state
    init_backlink_reclaim_state()

    # Check state AFTER upload section (in case new file was just processed)
    has_broken_links = st.session_state.df is not None
    has_redirect_chains = st.session_state.rc_df is not None
    has_image_alt_text = st.session_state.iat_df is not None
    has_backlink_reclaim = st.session_state.br_grouped_pages and len(st.session_state.br_grouped_pages) > 0
    has_post_ids = st.session_state.has_post_ids

    # If data is loaded, show the workflow below
    if has_broken_links or has_redirect_chains or has_image_alt_text or has_backlink_reclaim:
        st.markdown("---")

        # Show task switcher if multiple tasks
        render_task_switcher()

        # Render based on current task type
        current_task = st.session_state.current_task

        if current_task == 'redirect_chains' and has_redirect_chains:
            # Redirect Chains workflow
            st.markdown('<p class="section-header">🔄 Redirect Chains</p>', unsafe_allow_html=True)
            render_rc_metrics()
            st.markdown("---")
            render_rc_warnings()
            render_rc_spreadsheet()
            st.markdown("---")
            render_wordpress_section()
            st.markdown("---")
            render_rc_export_section()

        elif current_task == 'broken_links' and has_broken_links:
            # Broken Links workflow (existing)
            st.markdown('<p class="section-header">🔗 Broken Links</p>', unsafe_allow_html=True)
            render_metrics()
            st.markdown("---")
            render_spreadsheet()
            st.markdown("---")
            render_wordpress_section()
            st.markdown("---")
            render_export_section()

        elif current_task == 'image_alt_text' and has_image_alt_text:
            # Image Alt Text workflow
            st.markdown('<p class="section-header">🖼️ Image Alt Text</p>', unsafe_allow_html=True)
            render_iat_metrics()
            st.markdown("---")
            render_iat_spreadsheet()
            st.markdown("---")
            render_wordpress_section()
            st.markdown("---")
            render_iat_export_section()

        elif current_task == 'backlink_reclaim' and has_backlink_reclaim:
            # Backlink Reclaim workflow
            st.markdown('<p class="section-header">🔙 Backlink Reclaim</p>', unsafe_allow_html=True)
            # Get WordPress client if connected
            wp_client = st.session_state.wp_client if st.session_state.wp_connected else None
            render_backlink_fix_workflow(
                scan_results=st.session_state.br_scan_results,
                domain=st.session_state.br_domain,
                wp_client=wp_client
            )

        else:
            # Fallback - show whatever data we have
            if has_redirect_chains:
                st.session_state.current_task = 'redirect_chains'
                st.rerun()
            elif has_broken_links:
                st.session_state.current_task = 'broken_links'
                st.rerun()
            elif has_image_alt_text:
                st.session_state.current_task = 'image_alt_text'
                st.rerun()
            elif has_backlink_reclaim:
                st.session_state.current_task = 'backlink_reclaim'
                st.rerun()
    
    render_about()
    render_instructions()
    render_footer()


if __name__ == "__main__":
    # Check for landing page routing
    feature = st.query_params.get("feature", "")

    if feature == "reclaim":
        # Load the backlink reclaim landing page
        try:
            from landing_reclaim import render_landing_page
            render_landing_page()
        except ImportError as e:
            st.error(f"Landing page module not found: {e}")
            main()
    else:
        main()
